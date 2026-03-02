"""
excel_handler.py - Excel 模板生成与解析

功能：
  1. generate_template(): 生成一个 6-Sheet 的 .xlsx 模板文件，
     包含表头、字段说明行、示例数据和下拉验证。
  2. parse_excel(): 读取用户填好的 Excel，转换为标准 input_data.json 格式的 dict，
     并调用 validate_input() 做校验。
  3. compute_module_from_classification(): 根据「模块分类」Sheet + 子活动营收，
     自动聚合各模块营收，辅助填写 module_trend。

Sheet 列表：
  - Meta: 基础信息
  - 核心大盘趋势: 营收/ARPU/付费率时间序列
  - 模块营收趋势: 各模块营收时间序列
  - 用户分层ARPU: 超R/大R/中R ARPU
  - 子活动明细: 当期子活动诊断
  - 模块分类（可选）: 子活动 → 模块映射表

依赖：openpyxl
"""

import json
import os
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# 样式常量
# ============================================================

_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_HINT_FONT = Font(name="Microsoft YaHei", italic=True, size=9, color="808080")
_HINT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_HINT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_EXAMPLE_FONT = Font(name="Microsoft YaHei", size=10, color="666666")

_DATA_FONT = Font(name="Microsoft YaHei", size=10)
_DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ============================================================
# Sheet 定义
# ============================================================

# 每个 sheet 的配置：(sheet_name, headers, hints, example_rows, col_widths)
# headers: [(display_name, json_key), ...]

_SHEET_META = {
    "name": "Meta",
    "headers": [
        ("活动名称 (event_name)", "event_name"),
        ("对标活动 (benchmark_event)", "benchmark_event"),
    ],
    "hints": [
        "填写当期活动名称，如 2026春节",
        "填写同比对标活动名称，如 2025春节",
    ],
    "examples": [
        ["2026春节", "2025春节"],
    ],
    "col_widths": [30, 30],
}

_SHEET_METRICS = {
    "name": "核心大盘趋势",
    "headers": [
        ("活动名称 (event)", "event"),
        ("总营收 USD (revenue)", "revenue"),
        ("ARPU USD (arpu)", "arpu"),
        ("付费率 % (pay_rate)", "pay_rate"),
    ],
    "hints": [
        "按时间正序填写活动名",
        "该期活动总营收（美元）",
        "人均付费金额（美元）",
        "付费率，填数字如 32.66 表示 32.66%",
    ],
    "examples": [
        ["2025周年庆", 1668966.19, 273.83, 34.43],
        ["2025音乐节", 320158.61, 74.84, 27.43],
        ["2025万圣节", 613401.28, 144.43, 30.09],
        ["2025感恩节", 722558.47, 161.72, 32.48],
        ["2025圣诞节", 938756.60, 193.88, 33.05],
        ["2026春节", 890703.55, 198.42, 32.66],
        ["2025春节（对标）", 502022.91, 90.83, 26.10],
    ],
    "col_widths": [22, 20, 16, 18],
}

_SHEET_MODULE = {
    "name": "模块营收趋势",
    "headers": [
        ("活动名称 (event)", "event"),
        ("外显类 USD (appearance)", "appearance"),
        ("小游戏 USD (minigame)", "minigame"),
        ("混合/养成 USD (hybrid)", "hybrid"),
    ],
    "hints": [
        "与「核心大盘趋势」Sheet 的时间节点一一对应",
        "外显类模块营收（美元）",
        "小游戏模块营收（美元），若无填 0",
        "混合/养成类模块营收（美元）",
    ],
    "examples": [
        ["2025周年庆", 980112, 112213.45, 576640.74],
        ["2025音乐节", 80025, 52778.9, 187354.71],
        ["2025万圣节", 320913, 71658.63, 220829.65],
        ["2025感恩节", 285651, 140663.85, 296243.62],
        ["2025圣诞节", 266769, 276193.03, 395794.57],
        ["2026春节", 412540.31, 303807.97, 174355.27],
        ["2025春节（对标）", 351700, 0, 150322.91],
    ],
    "col_widths": [22, 22, 22, 22],
}

_SHEET_USER_TIER = {
    "name": "用户分层ARPU",
    "headers": [
        ("活动名称 (event)", "event"),
        ("超R ARPU (super_r)", "super_r"),
        ("大R ARPU (big_r)", "big_r"),
        ("中R ARPU (mid_r)", "mid_r"),
    ],
    "hints": [
        "至少填 2 行：当期 + 对标；可选填更多历史期",
        "超R用户人均付费（美元）",
        "大R用户人均付费（美元）",
        "中R用户人均付费（美元）",
    ],
    "examples": [
        ["2026春节", 542.04, 121.44, 24.95],
        ["2025圣诞节", 528.21, 134.31, 35.14],
        ["2025春节（对标）", 311.10, 73.70, 18.10],
    ],
    "col_widths": [22, 20, 20, 20],
}

_SHEET_SUB_ACTIVITY = {
    "name": "子活动明细",
    "headers": [
        ("子活动名称 (name)", "name"),
        ("类型 (type)", "type"),
        ("营收 USD (revenue)", "revenue"),
        ("诊断 (status)", "status"),
        ("诊断理由 (reason)", "reason"),
    ],
    "hints": [
        "填写子活动/玩法名称",
        "下拉选择：外显 / 小游戏 / 养成 / 混合",
        "该子活动营收（美元）",
        "下拉选择：Keep（保留）或 Optimize（待优化）",
        "简述诊断理由和建议",
    ],
    "examples": [
        ["限定皮肤礼包", "外显", 245000, "Keep", "高转化率，用户反馈积极，连续3期TOP1"],
        ["春节拼图大作战", "小游戏", 180000, "Keep", "新增玩法，参与率超预期"],
        ["年兽养成计划", "养成", 95000, "Optimize", "养成周期过长导致中途流失率高"],
        ["新春抽奖转盘", "混合", 120000, "Keep", "经典保留玩法，稳定贡献营收"],
    ],
    "col_widths": [22, 14, 18, 14, 50],
}

_SHEET_MODULE_CLASSIFY = {
    "name": "模块分类",
    "headers": [
        ("子活动/玩法名称 (name)", "name"),
        ("所属模块 (module)", "module"),
        ("备注 (note)", "note"),
    ],
    "hints": [
        "填写子活动名称（需与「子活动明细」Sheet 的名称一致）",
        "下拉选择：外显 / 小游戏 / 养成 / 混合",
        "可选，补充说明",
    ],
    "examples": [
        ["限定皮肤礼包", "外显", "行军皮肤套装等"],
        ["春节拼图大作战", "小游戏", ""],
        ["年兽养成计划", "养成", "养成 + 抢购礼包"],
        ["新春抽奖转盘", "混合", "GACHA + BP"],
        ["主城皮肤GACHA", "外显", "纯本体抽奖"],
        ["挖孔小游戏", "小游戏", "连续两期上线"],
        ["钓鱼小游戏", "小游戏", "新增玩法"],
        ["通行证BP", "混合", "长周期付费"],
    ],
    "col_widths": [28, 18, 40],
}

# 模块名称 → JSON 字段的映射（支持常见别名）
_MODULE_NAME_TO_KEY = {
    "外显": "appearance",
    "小游戏": "minigame",
    "养成": "hybrid",
    "混合": "hybrid",
}

# 用户可能填写的别名 → 标准名称
_MODULE_ALIAS = {
    "养成线": "养成",
    "外显类": "外显",
    "小游戏类": "小游戏",
    "混合类": "混合",
    "混合/养成": "混合",
}

_ALL_SHEETS = [
    _SHEET_META,
    _SHEET_METRICS,
    _SHEET_MODULE,
    _SHEET_USER_TIER,
    _SHEET_SUB_ACTIVITY,
    _SHEET_MODULE_CLASSIFY,
]


# ============================================================
# 模板生成
# ============================================================

def generate_template(output_path: str) -> str:
    """
    生成一个 6-Sheet 的 .xlsx 空白模板文件。

    每个 Sheet 结构：
      - 第 1 行：表头（蓝底白字加粗）
      - 第 2 行：字段说明/提示（灰色斜体）
      - 第 3 行起：示例数据（浅黄色背景灰色字体），用户可覆盖或删除

    Args:
        output_path: 输出文件路径，如 "report_images/template.xlsx"

    Returns:
        str: 生成的文件绝对路径
    """
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for sheet_cfg in _ALL_SHEETS:
        ws = wb.create_sheet(title=sheet_cfg["name"])
        headers = sheet_cfg["headers"]
        hints = sheet_cfg["hints"]
        examples = sheet_cfg["examples"]
        col_widths = sheet_cfg["col_widths"]

        num_cols = len(headers)

        # ── 第 1 行：表头 ──
        for col_idx, (display_name, _) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

        # ── 第 2 行：字段说明 ──
        for col_idx, hint_text in enumerate(hints, start=1):
            cell = ws.cell(row=2, column=col_idx, value=hint_text)
            cell.font = _HINT_FONT
            cell.fill = _HINT_FILL
            cell.alignment = _HINT_ALIGNMENT
            cell.border = _THIN_BORDER

        # ── 第 3 行起：示例数据 ──
        for row_offset, example_row in enumerate(examples):
            row_num = 3 + row_offset
            for col_idx, value in enumerate(example_row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = _EXAMPLE_FONT
                cell.fill = _EXAMPLE_FILL
                cell.alignment = _DATA_ALIGNMENT
                cell.border = _THIN_BORDER

        # ── 列宽 ──
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── 冻结首行 ──
        ws.freeze_panes = "A3"

        # ── 数据验证 ──
        if sheet_cfg["name"] == "子活动明细":
            # type 列下拉
            type_dv = DataValidation(
                type="list",
                formula1='"外显,小游戏,养成,混合"',
                allow_blank=True,
            )
            type_dv.error = "请从下拉列表中选择：外显 / 小游戏 / 养成 / 混合"
            type_dv.errorTitle = "类型错误"
            type_dv.prompt = "选择子活动类型"
            type_dv.promptTitle = "类型"
            type_col = get_column_letter(2)  # type 是第 2 列
            type_dv.add(f"{type_col}3:{type_col}100")
            ws.add_data_validation(type_dv)

            # status 列下拉
            status_dv = DataValidation(
                type="list",
                formula1='"Keep,Optimize"',
                allow_blank=True,
            )
            status_dv.error = "请从下拉列表中选择：Keep 或 Optimize"
            status_dv.errorTitle = "状态错误"
            status_dv.prompt = "Keep = 表现优秀保留；Optimize = 待优化"
            status_dv.promptTitle = "诊断状态"
            status_col = get_column_letter(4)  # status 是第 4 列
            status_dv.add(f"{status_col}3:{status_col}100")
            ws.add_data_validation(status_dv)

        elif sheet_cfg["name"] == "模块分类":
            # module 列下拉
            module_dv = DataValidation(
                type="list",
                formula1='"外显,小游戏,养成,混合"',
                allow_blank=True,
            )
            module_dv.error = "请从下拉列表中选择：外显 / 小游戏 / 养成 / 混合"
            module_dv.errorTitle = "模块错误"
            module_dv.prompt = "选择该子活动所属的模块分类"
            module_dv.promptTitle = "所属模块"
            module_col = get_column_letter(2)  # module 是第 2 列
            module_dv.add(f"{module_col}3:{module_col}100")
            ws.add_data_validation(module_dv)

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f"[excel_handler] 模板已生成: {abs_path}")
    return abs_path


# ============================================================
# Excel 解析
# ============================================================

def parse_excel(excel_path: str) -> dict:
    """
    读取用户填好的 Excel，转换为标准 input_data.json 格式。

    解析规则：
      - 第 1 行为表头（跳过）
      - 第 2 行为提示行（跳过）
      - 第 3 行起为数据行，跳过全空行

    Args:
        excel_path: 填好的 .xlsx 文件路径

    Returns:
        dict: 标准化数据，与 input_data.json 格式一致

    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 数据校验失败，包含详细中文错误信息
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    result = {}
    parse_errors = []

    # ── 解析 Meta ──
    result["meta"] = _parse_meta_sheet(wb, parse_errors)

    # ── 解析 核心大盘趋势 ──
    result["metrics_trend"] = _parse_table_sheet(
        wb, "核心大盘趋势",
        _SHEET_METRICS["headers"],
        {"revenue": float, "arpu": float, "pay_rate": float},
        parse_errors,
    )

    # ── 解析 模块营收趋势 ──
    result["module_trend"] = _parse_table_sheet(
        wb, "模块营收趋势",
        _SHEET_MODULE["headers"],
        {"appearance": float, "minigame": float, "hybrid": float},
        parse_errors,
    )

    # ── 解析 用户分层ARPU ──
    result["user_tier_trend"] = _parse_table_sheet(
        wb, "用户分层ARPU",
        _SHEET_USER_TIER["headers"],
        {"super_r": float, "big_r": float, "mid_r": float},
        parse_errors,
    )

    # ── 解析 子活动明细 ──
    result["sub_activity_detail"] = _parse_table_sheet(
        wb, "子活动明细",
        _SHEET_SUB_ACTIVITY["headers"],
        {"revenue": float},
        parse_errors,
    )

    # ── 解析 模块分类（必填 Sheet）──
    module_classify = _parse_module_classify_sheet(wb, parse_errors)
    result["module_classification"] = module_classify
    # 自动补全 sub_activity_detail 中缺失的 type 字段
    if module_classify:
        _enrich_sub_activity_types(result["sub_activity_detail"], module_classify)

    wb.close()

    # ── 检查解析阶段错误 ──
    if parse_errors:
        raise ValueError(
            "Excel 解析错误:\n" + "\n".join(f"  - {e}" for e in parse_errors)
        )

    # ── 调用 validate_input 做业务校验 ──
    try:
        from chart_generator import validate_input
    except ImportError:
        import sys
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        from chart_generator import validate_input

    validation_errors = validate_input(result)
    if validation_errors:
        raise ValueError(
            "数据校验失败:\n" + "\n".join(f"  - {e}" for e in validation_errors)
        )

    print(f"[excel_handler] Excel 解析成功，共解析:")
    print(f"  - metrics_trend: {len(result['metrics_trend'])} 条")
    print(f"  - module_trend: {len(result['module_trend'])} 条")
    print(f"  - user_tier_trend: {len(result['user_tier_trend'])} 条")
    print(f"  - sub_activity_detail: {len(result['sub_activity_detail'])} 条")
    print(f"  - module_classification: {len(result['module_classification'])} 条映射")

    # 自动计算当期模块聚合（供参考）
    if result["module_classification"]:
        agg = compute_module_from_classification(
            result["sub_activity_detail"],
            result["module_classification"],
        )
        print(f"  - 基于分类自动聚合当期模块营收:")
        print(f"      外显类 (appearance): ${agg['appearance']:,.0f}")
        print(f"      小游戏 (minigame):   ${agg['minigame']:,.0f}")
        print(f"      混合/养成 (hybrid):  ${agg['hybrid']:,.0f}")

    return result


def _parse_meta_sheet(wb: Any, errors: list) -> dict:
    """解析 Meta Sheet"""
    sheet_name = "Meta"
    if sheet_name not in wb.sheetnames:
        errors.append(f"找不到 Sheet「{sheet_name}」")
        return {"event_name": "", "benchmark_event": ""}

    ws = wb[sheet_name]
    meta = {"event_name": "", "benchmark_event": ""}

    # 第 3 行第 1 列 = event_name, 第 2 列 = benchmark_event
    row = 3
    val_1 = ws.cell(row=row, column=1).value
    val_2 = ws.cell(row=row, column=2).value

    if val_1 is None or str(val_1).strip() == "":
        errors.append(f"Sheet「{sheet_name}」第 {row} 行: 活动名称不能为空")
    else:
        meta["event_name"] = str(val_1).strip()

    if val_2 is None or str(val_2).strip() == "":
        errors.append(f"Sheet「{sheet_name}」第 {row} 行: 对标活动不能为空")
    else:
        meta["benchmark_event"] = str(val_2).strip()

    return meta


def _parse_table_sheet(
    wb: Any,
    sheet_name: str,
    headers: list[tuple[str, str]],
    numeric_fields: dict[str, type],
    errors: list,
) -> list[dict]:
    """
    通用解析：读取一个表格型 Sheet，返回 list[dict]。

    Args:
        wb: openpyxl Workbook
        sheet_name: Sheet 名称
        headers: [(display_name, json_key), ...]
        numeric_fields: {json_key: type} 需要转为数值的字段
        errors: 收集错误的列表
    """
    if sheet_name not in wb.sheetnames:
        errors.append(f"找不到 Sheet「{sheet_name}」")
        return []

    ws = wb[sheet_name]
    json_keys = [h[1] for h in headers]
    rows = []

    # 从第 3 行开始读取数据
    for row_idx in range(3, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(json_keys) + 1)]

        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        record = {}
        for col_idx, (json_key, value) in enumerate(zip(json_keys, row_values)):
            if json_key in numeric_fields:
                # 数值字段处理
                if value is None or str(value).strip() == "":
                    errors.append(
                        f"Sheet「{sheet_name}」第 {row_idx} 行「{headers[col_idx][0]}」: 数值不能为空"
                    )
                    record[json_key] = 0.0
                else:
                    try:
                        record[json_key] = float(value)
                    except (ValueError, TypeError):
                        errors.append(
                            f"Sheet「{sheet_name}」第 {row_idx} 行「{headers[col_idx][0]}」: "
                            f"无法转换为数字 (当前值: {value})"
                        )
                        record[json_key] = 0.0
            else:
                # 文本字段处理
                if value is None or str(value).strip() == "":
                    errors.append(
                        f"Sheet「{sheet_name}」第 {row_idx} 行「{headers[col_idx][0]}」: 不能为空"
                    )
                    record[json_key] = ""
                else:
                    record[json_key] = str(value).strip()

        rows.append(record)

    return rows


def _parse_module_classify_sheet(wb: Any, errors: list | None = None) -> list[dict]:
    """
    解析「模块分类」Sheet（必填）。

    返回 list[dict]，每项 {"name": "xxx", "module": "外显", "note": "..."}。
    如果 Sheet 不存在或为空，向 errors 中追加错误。
    """
    sheet_name = "模块分类"
    if sheet_name not in wb.sheetnames:
        if errors is not None:
            errors.append(f"找不到必填 Sheet「{sheet_name}」，请在模块分类表中填写每个子活动所属的模块")
        return []

    ws = wb[sheet_name]
    rows = []

    for row_idx in range(3, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=1).value
        module_val = ws.cell(row=row_idx, column=2).value
        note_val = ws.cell(row=row_idx, column=3).value

        # 跳过空行
        if name_val is None or str(name_val).strip() == "":
            continue

        name = str(name_val).strip()
        module = str(module_val).strip() if module_val else ""
        note = str(note_val).strip() if note_val else ""

        # 自动修正别名
        if module in _MODULE_ALIAS:
            corrected = _MODULE_ALIAS[module]
            print(f"[excel_handler] 自动修正: 「{module}」→「{corrected}」(Row {row_idx})")
            module = corrected

        if not module:
            if errors is not None:
                errors.append(
                    f"Sheet「{sheet_name}」第 {row_idx} 行「{name}」: 所属模块不能为空"
                )
        elif module not in _MODULE_NAME_TO_KEY:
            if errors is not None:
                errors.append(
                    f"Sheet「{sheet_name}」第 {row_idx} 行「{name}」: "
                    f"模块「{module}」不在已知列表中，请选择：外显 / 小游戏 / 养成 / 混合"
                )

        rows.append({"name": name, "module": module, "note": note})

    if not rows:
        if errors is not None:
            errors.append(f"Sheet「{sheet_name}」没有数据，请至少填写 1 条模块分类映射")
    else:
        print(f"[excel_handler] 模块分类: 读取 {len(rows)} 条映射")

    return rows


def _enrich_sub_activity_types(
    sub_activities: list[dict],
    module_classify: list[dict],
) -> None:
    """
    用模块分类表补全 sub_activity_detail 中的 type 字段。

    规则：
      - 如果子活动的 type 为空或未填，从模块分类表中查找同名活动的 module 并填入
      - 如果子活动已有 type，不覆盖
      - 匹配逻辑：精确匹配名称；如果找不到，尝试包含匹配
    """
    # 构建映射: name -> module
    classify_map = {}
    for item in module_classify:
        if item["name"] and item["module"]:
            classify_map[item["name"]] = item["module"]

    if not classify_map:
        return

    for activity in sub_activities:
        if activity.get("type"):
            continue  # 已有 type，不覆盖

        name = activity.get("name", "")

        # 精确匹配
        if name in classify_map:
            activity["type"] = classify_map[name]
            continue

        # 包含匹配（子活动名称包含分类表中的名称，或反过来）
        for cls_name, cls_module in classify_map.items():
            if cls_name in name or name in cls_name:
                activity["type"] = cls_module
                break


def compute_module_from_classification(
    sub_activities: list[dict],
    module_classify: list[dict] | None = None,
) -> dict[str, float]:
    """
    根据子活动明细 + 模块分类表，自动聚合当期各模块的营收。

    可用于辅助填写或校验 module_trend 中当期数据。

    Args:
        sub_activities: sub_activity_detail 列表
        module_classify: 模块分类表（可选，如果 sub_activities 已有 type 则不需要）

    Returns:
        dict: {"appearance": float, "minigame": float, "hybrid": float}
    """
    # 先用分类表补全 type
    if module_classify:
        _enrich_sub_activity_types(sub_activities, module_classify)

    aggregated = {"appearance": 0.0, "minigame": 0.0, "hybrid": 0.0}

    for activity in sub_activities:
        type_name = activity.get("type", "")
        revenue = float(activity.get("revenue", 0))
        module_key = _MODULE_NAME_TO_KEY.get(type_name)
        if module_key:
            aggregated[module_key] += revenue
        else:
            print(
                f"[excel_handler] 警告: 子活动「{activity.get('name', '?')}」"
                f"类型「{type_name}」无法映射到模块，营收未计入聚合"
            )

    return aggregated


def save_as_json(data: dict, output_path: str) -> str:
    """
    将解析后的数据保存为 JSON 文件。

    Args:
        data: parse_excel() 的返回值
        output_path: 输出路径

    Returns:
        str: 保存的文件绝对路径
    """
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    abs_path = os.path.abspath(output_path)
    print(f"[excel_handler] JSON 已保存: {abs_path}")
    return abs_path


# ============================================================
# CLI 入口
# ============================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Excel 模板生成与解析工具")
    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    # generate 子命令
    gen_parser = subparsers.add_parser("generate", help="生成空白 Excel 模板")
    gen_parser.add_argument(
        "--output", "-o",
        default="event_review_template.xlsx",
        help="输出文件路径 (默认: event_review_template.xlsx)",
    )

    # parse 子命令
    parse_parser = subparsers.add_parser("parse", help="解析填好的 Excel 为 JSON")
    parse_parser.add_argument("input", help="填好的 Excel 文件路径")
    parse_parser.add_argument(
        "--output", "-o",
        default=None,
        help="JSON 输出路径 (默认: 与输入同目录的 input_data.json)",
    )

    args = parser.parse_args()

    if args.command == "generate":
        generate_template(args.output)
    elif args.command == "parse":
        data = parse_excel(args.input)
        out_path = args.output or os.path.join(
            os.path.dirname(args.input) or ".", "input_data.json"
        )
        save_as_json(data, out_path)
    else:
        parser.print_help()
