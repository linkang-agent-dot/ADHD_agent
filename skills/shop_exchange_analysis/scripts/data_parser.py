"""
兑换商店数据解析模块
支持三种输入方式：文本粘贴、Excel 文件、JSON
"""

import re
import json
import os

# ============================================================
# 默认分类规则
# ============================================================
DEFAULT_CATEGORY_RULES = {
    "主城皮肤": ["皮肤"],
    "英雄养成": ["英雄", "升星", "碎片"],
    "军备养成": ["军备", "T6"],
    "装备养成": ["装备", "纳米", "晶体", "重铸"],
    "加速道具": ["加速"],
    "收藏品": ["收藏品"],
    "核心材料": ["机能核心"],
    "资源/抽奖": ["资源", "奖池", "宝箱", "抽奖"],
}

# 中文列名 → 标准字段名 映射
COLUMN_MAPPING = {
    "道具名称": "item_name",
    "兑换人次": "exchange_users",
    "兑换次数": "exchange_count",
    "人均兑换次数": "avg_exchanges_per_user",
    "平均消耗代币": "avg_token_cost",
    "代币价格": "token_price",
    "限购数量": "purchase_limit",
    "兑换饱和度": "saturation_rate",
    "类别": "category",
}


def parse_text_input(raw_text: str) -> list[dict]:
    """
    解析用户粘贴的文本表格数据。
    支持 Tab 分隔、| 分隔、多空格分隔。
    """
    lines = [l.strip() for l in raw_text.strip().split("\n") if l.strip()]
    if not lines:
        raise ValueError("输入文本为空")

    # 检测分隔符
    first_line = lines[0]
    if "\t" in first_line:
        sep = "\t"
    elif "|" in first_line:
        sep = "|"
    else:
        sep = "\t"  # 默认 Tab

    # 解析表头
    header_raw = [h.strip().strip("|") for h in lines[0].split(sep) if h.strip().strip("|")]
    header = []
    for h in header_raw:
        mapped = COLUMN_MAPPING.get(h, h)
        header.append(mapped)

    # 解析数据行
    items = []
    seen_names = {}
    for line in lines[1:]:
        # 跳过分隔线
        if re.match(r'^[\-\|=\s]+$', line):
            continue
        cells = [c.strip().strip("|") for c in line.split(sep) if c.strip().strip("|")]
        if len(cells) < len(header):
            # 补齐缺失列
            cells.extend([""] * (len(header) - len(cells)))

        row = {}
        for i, col_name in enumerate(header):
            if i < len(cells):
                val = cells[i].strip()
                # 去除 % 符号
                val = val.replace("%", "")
                row[col_name] = val

        # 转换数值类型
        item = _convert_types(row)
        if item and item.get("item_name"):
            # 处理重复道具名
            name = item["item_name"]
            if name in seen_names:
                seen_names[name] += 1
                suffix_map = {2: "②", 3: "③", 4: "④", 5: "⑤", 6: "⑥", 7: "⑦", 8: "⑧", 9: "⑨", 10: "⑩"}
                item["item_name"] = name + suffix_map.get(seen_names[name], f"({seen_names[name]})")
            else:
                seen_names[name] = 1
            items.append(item)

    if not items:
        raise ValueError("未能解析出有效的道具数据，请检查输入格式")

    # 自动补算缺失字段
    items = _fill_computed_fields(items)
    return items


def parse_json(data) -> list[dict]:
    """解析 JSON 数据（dict 或 str）"""
    if isinstance(data, str):
        data = json.loads(data)

    if isinstance(data, dict):
        items_raw = data.get("items", [])
        activity_name = data.get("activity_name", "")
    elif isinstance(data, list):
        items_raw = data
        activity_name = ""
    else:
        raise ValueError("JSON 数据格式不正确")

    items = []
    for row in items_raw:
        item = _convert_types(row)
        if item:
            items.append(item)

    items = _fill_computed_fields(items)
    return items


def parse_excel(filepath: str) -> list[dict]:
    """解析 Excel 文件"""
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("需要安装 pandas 和 openpyxl: pip install pandas openpyxl")

    df = pd.read_excel(filepath, engine='openpyxl')

    # 映射列名
    rename_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str in COLUMN_MAPPING:
            rename_map[col] = COLUMN_MAPPING[col_str]
        else:
            rename_map[col] = col_str
    df.rename(columns=rename_map, inplace=True)

    items = []
    seen_names = {}
    for _, row in df.iterrows():
        item = {}
        for col in df.columns:
            val = row[col]
            if pd.notna(val):
                item[col] = val
        item = _convert_types(item)
        if item and item.get("item_name"):
            name = item["item_name"]
            if name in seen_names:
                seen_names[name] += 1
                suffix_map = {2: "②", 3: "③", 4: "④", 5: "⑤"}
                item["item_name"] = name + suffix_map.get(seen_names[name], f"({seen_names[name]})")
            else:
                seen_names[name] = 1
            items.append(item)

    items = _fill_computed_fields(items)
    return items


def generate_excel_template(output_path: str):
    """生成 Excel 模板供用户填写"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "兑换数据"

    headers = ["道具名称", "兑换人次", "兑换次数", "人均兑换次数", "平均消耗代币",
               "代币价格", "限购数量", "兑换饱和度(%)", "类别(可选)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # 添加示例行
    example = ["万能英雄碎片-橙色", 72, 1139, 15.82, 9491.67, 600, 150, 10.55, ""]
    for i, v in enumerate(example, 1):
        ws.cell(row=2, column=i, value=v)

    # 调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 4, 12)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def auto_categorize(items: list[dict], custom_rules: dict = None) -> list[dict]:
    """
    自动为道具分类。
    已有 category 的不覆盖，未分类的按关键词匹配。
    """
    rules = {**DEFAULT_CATEGORY_RULES}
    if custom_rules:
        rules.update(custom_rules)

    for item in items:
        if item.get("category"):
            continue
        name = item.get("item_name", "")
        matched = False
        for category, keywords in rules.items():
            for kw in keywords:
                if kw in name:
                    item["category"] = category
                    matched = True
                    break
            if matched:
                break
        if not matched:
            item["category"] = "其他"

    return items


# ============================================================
# 内部工具函数
# ============================================================

def _convert_types(row: dict) -> dict:
    """将字符串值转换为正确的数值类型"""
    item = {}
    int_fields = {"exchange_users", "exchange_count", "token_price", "purchase_limit"}
    float_fields = {"avg_exchanges_per_user", "avg_token_cost", "saturation_rate"}
    str_fields = {"item_name", "category"}

    for k, v in row.items():
        if k in str_fields:
            item[k] = str(v).strip() if v else ""
        elif k in int_fields:
            try:
                item[k] = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                item[k] = 0
        elif k in float_fields:
            try:
                item[k] = round(float(str(v).replace(",", "")), 2)
            except (ValueError, TypeError):
                item[k] = 0.0
        else:
            item[k] = v

    return item


def _fill_computed_fields(items: list[dict]) -> list[dict]:
    """自动补算可推导字段"""
    for item in items:
        users = item.get("exchange_users", 0)
        count = item.get("exchange_count", 0)
        price = item.get("token_price", 0)
        limit = item.get("purchase_limit", 1)

        # 补算 avg_exchanges_per_user
        if not item.get("avg_exchanges_per_user") and users > 0:
            item["avg_exchanges_per_user"] = round(count / users, 2)

        # 补算 avg_token_cost
        if not item.get("avg_token_cost") and users > 0:
            avg_ex = item.get("avg_exchanges_per_user", 0)
            item["avg_token_cost"] = round(avg_ex * price, 2)

        # 补算 saturation_rate
        if not item.get("saturation_rate") and limit > 0:
            avg_ex = item.get("avg_exchanges_per_user", 0)
            item["saturation_rate"] = round(avg_ex / limit * 100, 2)

        # 补算 total_token_consumption（衍生字段）
        item["total_token_consumption"] = count * price

    return items


def validate_items(items: list[dict]) -> dict:
    """
    校验数据完整性和合法性。
    返回 {"valid": bool, "errors": list[str], "warnings": list[str]}
    """
    errors = []
    warnings = []

    if not items:
        errors.append("道具数据列表为空")
        return {"valid": False, "errors": errors, "warnings": warnings}

    for i, item in enumerate(items):
        prefix = f"第{i+1}条 [{item.get('item_name', '未知')}]"

        if not item.get("item_name"):
            errors.append(f"{prefix}: 缺少道具名称")
        if item.get("exchange_users", 0) < 0:
            errors.append(f"{prefix}: 兑换人次不能为负数")
        if item.get("exchange_count", 0) < 0:
            errors.append(f"{prefix}: 兑换次数不能为负数")
        if item.get("token_price", 0) <= 0:
            errors.append(f"{prefix}: 代币价格必须大于0")
        if item.get("purchase_limit", 0) <= 0:
            errors.append(f"{prefix}: 限购数量必须大于0")
        if item.get("saturation_rate", 0) > 200:
            warnings.append(f"{prefix}: 饱和度 {item['saturation_rate']}% 异常偏高，请确认数据")

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
    }
