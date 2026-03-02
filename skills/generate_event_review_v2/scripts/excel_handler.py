"""
Excel 模板生成 & 解析模块。

提供 8 个 Sheet 的 Excel 模板（Meta/触达转化/行为数据/付费整体/R级付费/付费转化/核心奖励/商业化礼包），
支持生成空白模板和解析填写后的 Excel 文件。
"""

import json
import os
import sys
import argparse
from typing import Dict, List, Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 样式常量
# ============================================================
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DESC_FONT = Font(name="Microsoft YaHei", italic=True, size=9, color="808080")
DESC_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
EXAMPLE_FONT = Font(name="Microsoft YaHei", size=10, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ============================================================
# Sheet 定义
# ============================================================
SHEET_DEFINITIONS = {
    "Meta": {
        "headers": ["字段名", "字段值"],
        "descriptions": ["填写说明", ""],
        "rows": [
            ["event_name（活动名称）*必填", "2026春节活动"],
            ["event_type（活动类型）", "节日活动"],
            ["event_start_date（开始日期）", "2026-01-25"],
            ["event_end_date（结束日期）", "2026-02-08"],
            ["change_description（本期改动）*必填", "本期新增了XX玩法，调整了YY概率"],
            ["benchmark_event（对标活动）", "2025春节活动"],
        ],
        "col_widths": [35, 50],
    },
    "触达转化": {
        "headers": ["阶段名称*", "当期人数*", "备注", "对标活动名称", "对标人数"],
        "descriptions": ["漏斗阶段名", "该阶段用户数", "可选备注", "对标活动（仅第1行填写）", "对标各阶段人数"],
        "rows": [
            ["推送触达", 1000000, "全服推送", "2025春节活动", 950000],
            ["弹窗展示", 800000, "", "", 750000],
            ["点击进入", 350000, "", "", 300000],
            ["活动参与", 200000, "至少完成1次任务", "", 180000],
            ["付费转化", 50000, "产生付费行为", "", 42000],
        ],
        "col_widths": [20, 18, 25, 20, 18],
    },
    "行为数据": {
        "headers": ["指标名称*", "当期值*", "对标值", "单位*"],
        "descriptions": ["行为指标名", "当期数值", "对标活动数值", "单位（人/次/%等）"],
        "rows": [
            ["日均DAU", 500000, 480000, "人"],
            ["活动参与率", 45.2, 42.0, "%"],
            ["人均参与次数", 8.5, 7.2, "次"],
            ["平均参与时长", 25.3, 22.1, "分钟"],
        ],
        "col_widths": [20, 18, 18, 12],
    },
    "付费整体趋势": {
        "headers": ["活动/月份名称*", "付费总额*", "付费率(%)*", "ARPU*", "ARPPU*", "是否同比对标"],
        "descriptions": ["按时间升序，至少6条", "付费总额", "付费率百分比", "ARPU", "ARPPU", "填\"是\"标记同比对标行"],
        "rows": [
            ["2025-08活动", 850000, 11.2, 45.6, 120.3, ""],
            ["2025-09活动", 920000, 12.0, 48.2, 125.8, ""],
            ["2025-10活动", 780000, 10.5, 40.1, 110.5, ""],
            ["2025-11活动", 950000, 12.5, 50.0, 130.2, ""],
            ["2025-12活动", 1020000, 13.0, 52.5, 135.0, ""],
            ["2026春节活动", 1150000, 14.2, 58.0, 142.5, ""],
            ["2025春节活动（同比）", 800000, 10.5, 42.0, 115.0, "是"],
        ],
        "col_widths": [25, 18, 15, 12, 12, 15],
    },
    "R级付费数据": {
        "headers": ["活动名称*", "R级名称*", "付费总额*", "付费率(%)*", "ARPU*", "ARPPU*"],
        "descriptions": ["每个活动每个R级一行，至少6个活动", "R级名称", "该R级付费总额", "付费率", "ARPU", "ARPPU"],
        "rows": [
            ["2025-08活动", "超R", 300000, 95.0, 5000, 5263],
            ["2025-08活动", "大R", 250000, 80.0, 1200, 1500],
            ["2025-08活动", "中R", 180000, 55.0, 300, 545],
            ["2025-08活动", "小R", 80000, 25.0, 50, 200],
            ["2025-08活动", "非R", 40000, 5.0, 5, 100],
        ],
        "col_widths": [20, 12, 18, 15, 12, 12],
    },
    "付费转化": {
        "headers": ["数据类型*", "价位(元)*", "购买笔数*", "购买人数*"],
        "descriptions": ["填\"当期\"或\"对标\"", "价位档", "购买笔数", "购买人数"],
        "rows": [
            ["当期", 6, 50000, 45000],
            ["当期", 30, 30000, 22000],
            ["当期", 68, 18000, 12000],
            ["当期", 128, 8000, 5000],
            ["当期", 328, 3000, 1800],
            ["当期", 648, 1500, 900],
            ["对标", 6, 48000, 43000],
            ["对标", 30, 28000, 20000],
        ],
        "col_widths": [15, 15, 18, 18],
    },
    "核心奖励": {
        "headers": ["奖励名称*", "预期产出*", "实际产出*", "单位*", "预期成本", "实际成本", "成本单位"],
        "descriptions": ["奖励名", "预期产出数量", "实际产出数量", "产出单位", "预期获取成本", "实际获取成本", "成本单位"],
        "rows": [
            ["SSR角色碎片", 100, 95, "个", 3280, 3450, "元"],
            ["限定皮肤", 1, 1, "件", 648, 680, "元"],
        ],
        "col_widths": [20, 15, 15, 10, 15, 15, 10],
    },
    "商业化礼包": {
        "headers": ["数据类型*", "礼包名称*", "价格*", "付费总额*", "付费人数*", "内容描述"],
        "descriptions": ["填\"当期\"或\"对标\"", "礼包名", "价格(元)", "付费总额", "付费人数", "礼包内容(可选)"],
        "rows": [
            ["当期", "新春超值礼包", 128, 640000, 5000, "含SSR碎片x10+金币x100000"],
            ["当期", "每日特惠礼包", 6, 300000, 50000, "含体力x120+抽奖券x5"],
            ["对标", "去年新春超值礼包", 128, 580000, 4500, ""],
        ],
        "col_widths": [12, 22, 12, 18, 15, 35],
    },
}


# ============================================================
# 模板生成
# ============================================================
def generate_template(output_path: str) -> str:
    """
    生成空白 Excel 模板，返回文件路径。

    Args:
        output_path: 输出文件路径 (.xlsx)

    Returns:
        生成的文件绝对路径
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb = Workbook()

    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, definition in SHEET_DEFINITIONS.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = definition["headers"]
        descriptions = definition["descriptions"]
        example_rows = definition["rows"]
        col_widths = definition["col_widths"]

        # 设置列宽
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # 第1行: 表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 第2行: 字段说明
        for col_idx, desc in enumerate(descriptions, 1):
            cell = ws.cell(row=2, column=col_idx, value=desc)
            cell.font = DESC_FONT
            cell.fill = DESC_FILL
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

        # 第3行起: 示例数据
        for row_idx, row_data in enumerate(example_rows, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = EXAMPLE_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER

        # 冻结前2行
        ws.freeze_panes = "A3"

    wb.save(output_path)
    return os.path.abspath(output_path)


# ============================================================
# Excel 解析
# ============================================================
def parse_excel(excel_path: str, skip_validation: bool = False) -> dict:
    """
    解析填好的 Excel 文件，返回标准 JSON 结构。

    自动处理:
    - 空 Sheet（如行为数据为空）不会阻断解析
    - Excel 中的 #DIV/0!、#N/A 等错误值自动转为 0
    - ARPU/ARPPU 为 0 时，自动从触达数据（参与人数/付费人数）和付费总额补算

    Args:
        excel_path: Excel 文件路径
        skip_validation: 是否跳过数据校验（默认 False）

    Returns:
        标准化数据字典

    Raises:
        ValueError: 数据格式错误时抛出中文错误信息（仅在 skip_validation=False 时）
    """
    wb = load_workbook(excel_path, data_only=True)
    data = {}

    data["meta"] = _parse_meta(wb)
    data["reach_conversion"] = _parse_reach_conversion(wb)
    data["behavior_data"] = _parse_behavior_data(wb)
    data["payment_overview"] = _parse_payment_overview(wb)
    data["r_tier_payment"] = _parse_r_tier_payment(wb)
    data["payment_conversion"] = _parse_payment_conversion(wb)
    data["core_reward"] = _parse_core_reward(wb)
    data["gift_packages"] = _parse_gift_packages(wb)

    # 自动补算: 用触达数据修复 ARPU/ARPPU 为 0 的记录（Excel 中 #DIV/0! 等公式错误）
    _auto_fix_arpu_arppu(data)

    # 校验
    if not skip_validation:
        from data_validator import validate_input
        result = validate_input(data)
        if not result["valid"]:
            raise ValueError("Excel 数据校验失败:\n" + "\n".join(result["errors"]))
        for w in result.get("warnings", []):
            print(f"  [WARN] {w}")

    return data


def _auto_fix_arpu_arppu(data: dict):
    """
    自动补算 payment_overview 中 ARPU/ARPPU 为 0 的记录。

    利用 reach_conversion 中的参与人数和付费人数，结合 revenue 反算。
    常见于 Excel 中公式结果为 #DIV/0! 的情况。
    """
    rc_stages = data.get("reach_conversion", {}).get("stages", [])
    po_ts = data.get("payment_overview", {}).get("time_series", [])
    if not rc_stages or not po_ts:
        return

    participate_users = 0
    paying_users = 0
    for s in rc_stages:
        stage = s.get("stage", "")
        if "参与" in stage:
            participate_users = s.get("users", 0)
        if "付费" in stage:
            paying_users = s.get("users", 0)

    if participate_users <= 0 and paying_users <= 0:
        return

    for item in po_ts:
        revenue = item.get("revenue", 0)
        if revenue > 0 and item.get("arpu", 0) == 0 and item.get("arppu", 0) == 0:
            if participate_users > 0:
                item["arpu"] = round(revenue / participate_users, 3)
            if paying_users > 0:
                item["arppu"] = round(revenue / paying_users, 3)
            print(f"  [AUTO-FIX] {item.get('event', '?')}: ARPU={item.get('arpu', 0):.3f}, ARPPU={item.get('arppu', 0):.3f}")


def _get_rows(wb, sheet_name: str, start_row: int = 3) -> List[List[Any]]:
    """读取指定 Sheet 从 start_row 开始的所有非空行"""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中未找到 Sheet: {sheet_name}")
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # 跳过全空行
        if all(v is None for v in row):
            continue
        rows.append(list(row))
    return rows


def _get_header_map(wb, sheet_name: str) -> Dict[str, int]:
    """读取 Sheet 第1行表头，返回 {关键词: 列索引} 映射。支持模糊匹配。"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(v).strip().lower() if v else "" for v in row]
    return {h: i for i, h in enumerate(headers) if h}


def _parse_meta(wb) -> dict:
    """解析 Meta Sheet"""
    rows = _get_rows(wb, "Meta")
    meta = {}
    field_map = {
        "event_name": "event_name",
        "event_type": "event_type",
        "event_start_date": "event_start_date",
        "event_end_date": "event_end_date",
        "change_description": "change_description",
        "benchmark_event": "benchmark_event",
    }
    for row in rows:
        if len(row) < 2 or row[0] is None:
            continue
        field_label = str(row[0]).strip()
        value = row[1]
        for key in field_map:
            if key in field_label.lower().replace("（", "(").replace("）", ")"):
                if value is not None:
                    meta[field_map[key]] = str(value).strip()
                break
    return meta


def _parse_reach_conversion(wb) -> dict:
    """解析触达转化 Sheet，支持多对标（额外列对: 对标活动名称-X, 对标人数-X）"""
    rows = _get_rows(wb, "触达转化")
    stages = []

    # 检测有多少组对标列（从第4列开始，每2列一组：活动名称 + 人数）
    # 格式: col3=对标活动名称, col4=对标人数, col5=对标活动名称2, col6=对标人数2, ...
    comparisons_data = {}  # {benchmark_name: [stage_data]}

    for row in rows:
        if len(row) < 2:
            continue
        stage_name = row[0]
        users = row[1] if len(row) > 1 else None
        note = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if stage_name and users is not None:
            user_count = _to_number(users)
            # 跳过用户数为0的阶段（表示该活动不涉及此环节）
            if user_count <= 0:
                continue
            stages.append({
                "stage": str(stage_name).strip(),
                "users": user_count,
                "note": note,
            })

        # 对标数据：扫描从 col3 开始的每一对列 (名称, 人数)
        col_idx = 3
        while col_idx + 1 < len(row):
            comp_name_val = row[col_idx]
            comp_users_val = row[col_idx + 1]

            if comp_name_val:
                comp_event = str(comp_name_val).strip()
                if comp_event not in comparisons_data:
                    comparisons_data[comp_event] = []

            # 找到最近设定的对标名（可能只在第一行填写）
            # 遍历已有的 comparisons_data 找到对应列位置的名称
            comp_event = None
            for row_check in rows:
                if len(row_check) > col_idx and row_check[col_idx]:
                    comp_event = str(row_check[col_idx]).strip()
                    break

            if comp_event and comp_users_val is not None and stage_name:
                comp_count = _to_number(comp_users_val)
                if comp_count > 0:
                    if comp_event not in comparisons_data:
                        comparisons_data[comp_event] = []
                    comparisons_data[comp_event].append({
                        "stage": str(stage_name).strip(),
                        "users": comp_count,
                    })

            col_idx += 2

    result = {"stages": stages}
    # 多对标数组
    if comparisons_data:
        result["comparisons"] = [
            {"benchmark_event": name, "stages": comp_stages}
            for name, comp_stages in comparisons_data.items()
            if comp_stages
        ]
        # 向后兼容
        if result["comparisons"]:
            result["comparison"] = result["comparisons"][0]
    return result


def _parse_behavior_data(wb) -> dict:
    """解析行为数据 Sheet"""
    rows = _get_rows(wb, "行为数据")
    metrics = []
    for row in rows:
        if len(row) < 4 or row[0] is None:
            continue
        m = {
            "metric_name": str(row[0]).strip(),
            "current_value": _to_number(row[1]),
            "unit": str(row[3]).strip() if row[3] else "",
        }
        if row[2] is not None:
            m["benchmark_value"] = _to_number(row[2])
        metrics.append(m)
    return {"metrics": metrics}


def _parse_payment_overview(wb) -> dict:
    """解析付费整体趋势 Sheet，根据表头动态匹配列顺序"""
    rows = _get_rows(wb, "付费整体趋势")
    hmap = _get_header_map(wb, "付费整体趋势")
    time_series = []
    yoy_benchmarks = {}  # {benchmark_name: entry}

    # 动态查找列索引（通过关键词匹配）
    def _find_col(keywords, default_idx):
        for k in keywords:
            for h, idx in hmap.items():
                if k in h:
                    return idx
        return default_idx

    col_event = _find_col(["活动", "月份", "名称"], 0)
    col_revenue = _find_col(["付费总额", "总额", "revenue"], 1)
    col_pay_rate = _find_col(["付费率", "pay_rate"], 2)
    # 区分 ARPU 和 ARPPU：先找 ARPPU，再找 ARPU（排除已匹配的列）
    col_arppu = _find_col(["arppu"], 4)
    col_arpu = None
    for h, idx in hmap.items():
        if "arpu" in h and "arppu" not in h and idx != col_arppu:
            col_arpu = idx
            break
    if col_arpu is None:
        col_arpu = 3
    col_yoy = _find_col(["同比", "yoy", "对标"], 5)

    for row in rows:
        if len(row) < 5 or row[col_event] is None:
            continue
        yoy_val = row[col_yoy] if len(row) > col_yoy else None
        is_benchmark = _is_benchmark_tag(yoy_val)
        entry = {
            "event": str(row[col_event]).strip(),
            "revenue": _to_number(row[col_revenue]),
            "pay_rate": _to_number(row[col_pay_rate]),
            "arpu": _to_number(row[col_arpu]),
            "arppu": _to_number(row[col_arppu]),
        }
        if is_benchmark:
            bench_name = _extract_benchmark_name(yoy_val, entry["event"])
            yoy_benchmarks[bench_name] = entry
        else:
            time_series.append(entry)

    result = {"time_series": time_series}
    # 多对标数组
    if yoy_benchmarks:
        result["yoy_benchmarks"] = list(yoy_benchmarks.values())
        # 向后兼容：保留单个 yoy_benchmark（取第一个）
        result["yoy_benchmark"] = list(yoy_benchmarks.values())[0]
    return result


def _parse_r_tier_payment(wb) -> dict:
    """解析 R 级付费数据 Sheet，根据表头动态匹配列顺序，支持多对标"""
    rows = _get_rows(wb, "R级付费数据")
    hmap = _get_header_map(wb, "R级付费数据")
    tiers_set = set()
    events_data = {}  # {event: {tier: {revenue, pay_rate, arpu, arppu}}}
    benchmark_events = {}  # {event: benchmark_name}

    # 动态查找列索引
    def _find_col(keywords, default_idx):
        for k in keywords:
            for h, idx in hmap.items():
                if k in h:
                    return idx
        return default_idx

    col_event = 0
    col_tier = 1
    col_revenue = _find_col(["付费总额", "总额", "revenue"], 2)
    col_pay_rate = _find_col(["付费率", "pay_rate"], 3)
    col_arppu = _find_col(["arppu"], 5)
    col_arpu = None
    for h, idx in hmap.items():
        if "arpu" in h and "arppu" not in h and idx != col_arppu:
            col_arpu = idx
            break
    if col_arpu is None:
        col_arpu = 4
    # 可选的"是否同比对标"列
    col_yoy = _find_col(["同比", "yoy", "对标"], None)

    for row in rows:
        if len(row) < 6 or row[col_event] is None or row[col_tier] is None:
            continue
        event = str(row[col_event]).strip()
        tier = str(row[col_tier]).strip()
        tiers_set.add(tier)

        # 检查对标标记
        if col_yoy is not None and len(row) > col_yoy:
            yoy_val = row[col_yoy]
            if _is_benchmark_tag(yoy_val) and event not in benchmark_events:
                benchmark_events[event] = _extract_benchmark_name(yoy_val, event)

        if event not in events_data:
            events_data[event] = {}
        events_data[event][tier] = {
            "revenue": _to_number(row[col_revenue]),
            "pay_rate": _to_number(row[col_pay_rate]),
            "arpu": _to_number(row[col_arpu]),
            "arppu": _to_number(row[col_arppu]),
        }

    # 保持 R 级顺序
    tier_order = ["超R", "大R", "中R", "小R", "非R"]
    tiers = [t for t in tier_order if t in tiers_set]
    tiers += [t for t in tiers_set if t not in tier_order]

    # 分离 time_series 和 benchmarks
    time_series = []
    benchmarks = []
    for event, data in events_data.items():
        entry = {"event": event, "data": data}
        if event in benchmark_events:
            entry["benchmark_name"] = benchmark_events[event]
            benchmarks.append(entry)
        else:
            time_series.append(entry)

    result = {"tiers": tiers, "time_series": time_series}
    if benchmarks:
        result["benchmarks"] = benchmarks
    return result


def _parse_payment_conversion(wb) -> dict:
    """解析付费转化 Sheet，支持 3 列（无购买人数）和 4 列格式，支持多对标"""
    rows = _get_rows(wb, "付费转化")
    current_tiers = []
    comparison_groups = {}  # {benchmark_name: [tiers]}

    for row in rows:
        if len(row) < 3 or row[0] is None:
            continue
        data_type = str(row[0]).strip()
        purchases = _to_number(row[2])
        # 兼容3列格式：无payers时默认等于purchases
        payers = _to_number(row[3]) if len(row) > 3 and row[3] is not None else purchases
        tier = {
            "price": _to_number(row[1]),
            "purchases": purchases,
            "payers": payers,
        }
        if "当期" in data_type or "current" in data_type.lower():
            current_tiers.append(tier)
        elif "对标" in data_type or "comparison" in data_type.lower():
            bench_name = _extract_benchmark_name(data_type, "对标")
            comparison_groups.setdefault(bench_name, []).append(tier)

    result = {
        "current": {
            "event": "",
            "price_tiers": current_tiers,
        }
    }
    # 多对标数组
    if comparison_groups:
        result["comparisons"] = [
            {"event": name, "price_tiers": tiers}
            for name, tiers in comparison_groups.items()
        ]
        # 向后兼容：保留单个 comparison（取第一个）
        first_name, first_tiers = next(iter(comparison_groups.items()))
        result["comparison"] = {"event": first_name, "price_tiers": first_tiers}
    return result


def _parse_core_reward(wb) -> dict:
    """解析核心奖励 Sheet"""
    rows = _get_rows(wb, "核心奖励")
    items = []
    for row in rows:
        if len(row) < 4 or row[0] is None:
            continue
        item = {
            "reward_name": str(row[0]).strip(),
            "expected_value": _to_number(row[1]),
            "actual_value": _to_number(row[2]),
            "unit": str(row[3]).strip() if row[3] else "",
        }
        if len(row) > 4 and row[4] is not None:
            item["expected_cost"] = _to_number(row[4])
        if len(row) > 5 and row[5] is not None:
            item["actual_cost"] = _to_number(row[5])
        if len(row) > 6 and row[6] is not None:
            item["cost_unit"] = str(row[6]).strip()
        items.append(item)
    return {"items": items}


def _parse_gift_packages(wb) -> dict:
    """解析商业化礼包 Sheet，根据表头动态匹配列顺序"""
    rows = _get_rows(wb, "商业化礼包")
    hmap = _get_header_map(wb, "商业化礼包")
    packages = []
    comparison_groups = {}  # {benchmark_name: [packages]}

    # 动态查找列索引
    def _find_col(keywords, default_idx):
        for k in keywords:
            for h, idx in hmap.items():
                if k in h:
                    return idx
        return default_idx

    col_type = 0
    col_name = _find_col(["礼包名", "名称"], 1)
    col_price = _find_col(["价格"], 2)
    col_revenue = _find_col(["付费总额", "总额", "revenue"], 4)
    col_payers = _find_col(["付费人数", "人数", "payers"], 3)
    col_desc = _find_col(["内容", "描述", "description"], 5)

    for row in rows:
        if len(row) < 5 or row[col_type] is None:
            continue
        data_type = str(row[col_type]).strip()
        pkg = {
            "package_name": str(row[col_name]).strip() if row[col_name] else "",
            "price": _to_number(row[col_price]),
            "revenue": _to_number(row[col_revenue]),
            "payers": _to_number(row[col_payers]),
        }
        if len(row) > col_desc and row[col_desc]:
            pkg["description"] = str(row[col_desc]).strip()

        if "当期" in data_type or "current" in data_type.lower():
            packages.append(pkg)
        elif "对标" in data_type or "comparison" in data_type.lower():
            bench_name = _extract_benchmark_name(data_type, "对标")
            comparison_groups.setdefault(bench_name, []).append(pkg)

    result = {"packages": packages}
    # 多对标数组
    if comparison_groups:
        result["comparisons"] = [
            {"event": name, "packages": pkgs}
            for name, pkgs in comparison_groups.items()
        ]
        # 向后兼容
        first_pkgs = next(iter(comparison_groups.values()))
        result["comparison_packages"] = first_pkgs
    return result


# ============================================================
# 工具函数
# ============================================================

def _is_benchmark_tag(value) -> bool:
    """判断一个字段值是否表示对标标记"""
    if value is None:
        return False
    tag = str(value).strip().lower()
    return tag in ("是", "yes", "true", "1") or "对标" in tag or "benchmark" in tag


def _extract_benchmark_name(tag_value, fallback_event: str = "") -> str:
    """
    从标记值中提取对标活动名称。

    支持格式：
    - "是" / "yes" / "true" / "1"  → 使用 fallback_event
    - "对标"                        → 使用 fallback_event 或 "对标"
    - "对标-2025春节"               → "2025春节"
    - "对标活动-2025圣诞节"         → "2025圣诞节"
    - "对标活动2025春节"            → "2025春节"（无连字符亦可）
    """
    if tag_value is None:
        return fallback_event or "对标"
    tag = str(tag_value).strip()

    # 布尔标记
    if tag.lower() in ("是", "yes", "true", "1"):
        return fallback_event or "对标"

    # "对标活动-XXX" / "对标-XXX" / "对标XXX" 前缀提取
    for prefix in ["对标活动-", "对标活动", "对标-", "对标"]:
        if tag.startswith(prefix):
            name = tag[len(prefix):].strip(" -")
            return name if name else (fallback_event or "对标")

    return fallback_event or tag


def _to_number(value) -> float:
    """将值转换为数值"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return 0


def save_as_json(data: dict, output_path: str) -> str:
    """保存数据为 JSON 文件，返回文件路径"""
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return os.path.abspath(output_path)


# ============================================================
# 命令行入口
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel 模板生成 & 解析")
    subparsers = parser.add_subparsers(dest="command")

    gen_parser = subparsers.add_parser("generate", help="生成空白 Excel 模板")
    gen_parser.add_argument("-o", "--output", required=True, help="输出文件路径")

    parse_parser = subparsers.add_parser("parse", help="解析填写后的 Excel")
    parse_parser.add_argument("-i", "--input", required=True, help="Excel 文件路径")
    parse_parser.add_argument("-o", "--output", required=True, help="输出 JSON 文件路径")

    args = parser.parse_args()

    if args.command == "generate":
        path = generate_template(args.output)
        print(f"模板已生成: {path}")
    elif args.command == "parse":
        data = parse_excel(args.input)
        save_as_json(data, args.output)
        print(f"解析完成，JSON 已保存: {os.path.abspath(args.output)}")
    else:
        parser.print_help()
