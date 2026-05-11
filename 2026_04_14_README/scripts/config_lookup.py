#!/usr/bin/env python3
"""X3 配置表查询工具 - 根据关键词在 xlsx 配置表中搜索"""

import sys
import os
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ⚠️ 线上问题排查必须读 master 环境配置（data_master），不要用 data_dev
# 其他人使用时，只需改这个路径为自己本地 SVN checkout 的 X3 配置表 master 目录
CONFIG_DIR = Path("/Users/zouhanling/Desktop/X3/design/data_master")


def list_files():
    """列出所有配置文件"""
    files = sorted(CONFIG_DIR.glob("*.xlsx"))
    for f in files:
        print(f.name)
    print(f"\n共 {len(files)} 个配置文件")


def list_sheets(filename):
    """列出指定文件的所有 sheet"""
    filepath = CONFIG_DIR / filename
    if not filepath.exists():
        print(f"ERROR: 文件不存在: {filename}", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        print(f"  {name} ({rows} rows x {cols} cols)")
    wb.close()


def read_sheet(filename, sheet_name=None, max_rows=50, search=None, columns=None):
    """读取指定 sheet 的数据
    
    Args:
        filename: xlsx 文件名
        sheet_name: sheet 名，不指定则读第一个
        max_rows: 最大返回行数
        search: 搜索关键词（在所有列中搜索）
        columns: 只返回指定列名（逗号分隔）
    """
    filepath = CONFIG_DIR / filename
    if not filepath.exists():
        print(f"ERROR: 文件不存在: {filename}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet 不存在: {sheet_name}", file=sys.stderr)
            print(f"可用 sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
            wb.close()
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]

    # 读取表头
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(row)]
        break

    if not headers:
        print(f"Sheet {sheet_name} 为空")
        wb.close()
        return

    # 过滤列
    col_indices = list(range(len(headers)))
    if columns:
        target_cols = [c.strip() for c in columns.split(",")]
        col_indices = [i for i, h in enumerate(headers) if h in target_cols]
        if not col_indices:
            print(f"WARNING: 未找到指定列，显示所有列", file=sys.stderr)
            col_indices = list(range(len(headers)))

    # 读取数据
    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        row_data = list(row)
        
        # 搜索过滤
        if search:
            search_lower = search.lower()
            matched = False
            for cell in row_data:
                if cell is not None and search_lower in str(cell).lower():
                    matched = True
                    break
            if not matched:
                continue

        # 构建结果行
        result = {}
        for i in col_indices:
            if i < len(row_data):
                val = row_data[i]
                if val is not None:
                    result[headers[i]] = val
        
        if result:
            result["_row"] = row_idx
            results.append(result)
        
        if len(results) >= max_rows:
            break

    wb.close()

    # 输出
    print(f"=== {filename} / {sheet_name} ===")
    print(f"表头: {', '.join(headers[i] for i in col_indices)}")
    print(f"结果: {len(results)} 行" + (f" (搜索: {search})" if search else ""))
    print()
    
    for r in results:
        row_num = r.pop("_row", "?")
        print(f"[Row {row_num}]")
        for k, v in r.items():
            print(f"  {k}: {v}")
        print()


def search_all(keyword, file_pattern=None, max_per_file=5):
    """在所有配置文件中搜索关键词
    
    Args:
        keyword: 搜索关键词
        file_pattern: 文件名过滤（部分匹配）
        max_per_file: 每个文件最多返回几条
    """
    files = sorted(CONFIG_DIR.glob("*.xlsx"))
    if file_pattern:
        files = [f for f in files if file_pattern.lower() in f.name.lower()]
    
    keyword_lower = keyword.lower()
    total_found = 0
    
    for filepath in files:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            continue
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            found_in_sheet = 0
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(row)]
                    continue
                
                for col_idx, cell in enumerate(row):
                    if cell is not None and keyword_lower in str(cell).lower():
                        if found_in_sheet == 0:
                            print(f"\n📄 {filepath.name} / {sheet_name}")
                        
                        col_name = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                        # 打印该行关键列
                        row_data = list(row)
                        summary = []
                        for i, h in enumerate(headers[:8]):  # 最多显示前8列
                            if i < len(row_data) and row_data[i] is not None:
                                summary.append(f"{h}={row_data[i]}")
                        print(f"  [Row {row_idx}] 匹配列={col_name} | {' | '.join(summary)}")
                        
                        found_in_sheet += 1
                        total_found += 1
                        break  # 一行只报一次
                
                if found_in_sheet >= max_per_file:
                    print(f"  ... (更多结果省略)")
                    break
        
        wb.close()
    
    print(f"\n共找到 {total_found} 条匹配结果")


def main():
    parser = argparse.ArgumentParser(description="X3 配置表查询工具")
    sub = parser.add_subparsers(dest="cmd")

    sub.add_parser("files", help="列出所有配置文件")

    p_sheets = sub.add_parser("sheets", help="列出文件的所有 sheet")
    p_sheets.add_argument("filename", help="xlsx 文件名")

    p_read = sub.add_parser("read", help="读取 sheet 数据")
    p_read.add_argument("filename", help="xlsx 文件名")
    p_read.add_argument("-s", "--sheet", help="sheet 名")
    p_read.add_argument("-n", "--max-rows", type=int, default=50, help="最大行数")
    p_read.add_argument("-q", "--search", help="搜索关键词")
    p_read.add_argument("-c", "--columns", help="指定列名(逗号分隔)")

    p_search = sub.add_parser("search", help="全局搜索")
    p_search.add_argument("keyword", help="搜索关键词")
    p_search.add_argument("-f", "--file-pattern", help="文件名过滤")
    p_search.add_argument("-n", "--max-per-file", type=int, default=5, help="每文件最多结果数")

    args = parser.parse_args()

    if args.cmd == "files":
        list_files()
    elif args.cmd == "sheets":
        list_sheets(args.filename)
    elif args.cmd == "read":
        read_sheet(args.filename, args.sheet, args.max_rows, args.search, args.columns)
    elif args.cmd == "search":
        search_all(args.keyword, args.file_pattern, args.max_per_file)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
