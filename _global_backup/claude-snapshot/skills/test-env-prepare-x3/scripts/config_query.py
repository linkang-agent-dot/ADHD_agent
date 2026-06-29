#!/usr/bin/env python3
"""Query X3 config tables (gdconfig xlsx) for item/activity IDs."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)

GDCONFIG = Path(__file__).resolve().parents[2] / ".." / ".." / "gdconfig" / "data"


def _find_gdconfig() -> Path:
    # Try relative to script, then cwd
    for base in [GDCONFIG, Path("gdconfig/data"), Path("../gdconfig/data")]:
        if base.exists():
            return base
    raise FileNotFoundError("gdconfig/data not found")


def _load_table(path: Path) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find english field name row (contains 'ID' in first column)
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "ID":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"No header row with 'ID' found in {path.name}")

    headers = [str(h) if h else f"col{i}" for i, h in enumerate(rows[header_idx])]
    data = []
    for row in rows[header_idx + 1:]:
        if row[0]:
            data.append(list(row))
    return headers, data


def query_item(keyword: str = None, item_id: str = None):
    base = _find_gdconfig()
    headers, data = _load_table(base / "Item.xlsx")

    id_idx = headers.index("ID")
    name_idx = headers.index("Name") if "Name" in headers else 1

    results = []
    for row in data:
        if item_id and str(row[id_idx]) == item_id:
            results.append(row)
        elif keyword and row[name_idx] and keyword in str(row[name_idx]):
            results.append(row)

    print(f"{'ID':<10} {'Name':<20} {'Type':<6} {'SubType':<8}")
    print("-" * 50)
    type_idx = headers.index("Type") if "Type" in headers else 4
    sub_idx = headers.index("SubType") if "SubType" in headers else 5
    for row in results[:30]:
        print(f"{row[id_idx]:<10} {str(row[name_idx] or ''):<20} {str(row[type_idx] or ''):<6} {str(row[sub_idx] or ''):<8}")
    if len(results) > 30:
        print(f"... and {len(results) - 30} more")


def query_activity(keyword: str = None, actv_id: str = None):
    base = _find_gdconfig()
    headers, data = _load_table(base / "ActvOnline.xlsx")

    id_idx = headers.index("ID")
    name_idx = headers.index("ActvName") if "ActvName" in headers else 2
    type_idx = headers.index("ActvType") if "ActvType" in headers else 5
    content_idx = headers.index("ContentID") if "ContentID" in headers else 4

    results = []
    for row in data:
        if actv_id and str(row[id_idx]) == actv_id:
            results.append(row)
        elif keyword:
            searchable = f"{row[name_idx] or ''} {row[1] or ''}"
            if keyword in searchable:
                results.append(row)

    print(f"{'ID':<10} {'Name':<20} {'Type':<6} {'ContentID':<10}")
    print("-" * 50)
    for row in results[:30]:
        print(f"{row[id_idx]:<10} {str(row[name_idx] or row[1] or ''):<20} {str(row[type_idx] or ''):<6} {str(row[content_idx] or ''):<10}")
    if len(results) > 30:
        print(f"... and {len(results) - 30} more")


def main():
    parser = argparse.ArgumentParser(description="Query X3 config tables")
    sub = parser.add_subparsers(dest="table")

    item_p = sub.add_parser("item", help="Query Item table")
    item_p.add_argument("--id", help="Exact item ID")
    item_p.add_argument("--name", help="Name keyword search")

    actv_p = sub.add_parser("activity", help="Query ActvOnline table")
    actv_p.add_argument("--id", help="Exact activity ID")
    actv_p.add_argument("--name", help="Name keyword search")

    args = parser.parse_args()
    if args.table == "item":
        query_item(keyword=args.name, item_id=args.id)
    elif args.table == "activity":
        query_activity(keyword=args.name, actv_id=args.id)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
