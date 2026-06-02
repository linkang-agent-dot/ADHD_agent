# -*- coding: utf-8 -*-
import io
try:
    import openpyxl.worksheet.table as tbl_mod
    from openpyxl.descriptors.base import String
    tbl_mod.Table.ref = String(allow_none=False)
except Exception: pass
import openpyxl
wb=openpyxl.load_workbook(r"C:\Users\linkang\Downloads\H_航海之路.xlsx", data_only=True, read_only=True)
out=io.open(r"C:\ADHD_agent\_tmp_scripts\xlsx.txt",'w',encoding='utf-8')
out.write("=== SHEETS ===\n")
for s in wb.sheetnames: out.write(f"  {s}\n")
sn=wb.sheetnames[0]
ws=wb[sn]
out.write(f"\n===== [{sn}] (前90非空行) =====\n")
cnt=0
for r in ws.iter_rows(values_only=True):
    cells=[str(c) for c in r if c is not None and str(c).strip()]
    if cells:
        out.write(" | ".join(cells)[:200]+"\n"); cnt+=1
    if cnt>=90: break
out.close(); print("sheets:",len(wb.sheetnames))
