# -*- coding: utf-8 -*-
import io
try:
    import openpyxl.worksheet.table as tbl_mod
    from openpyxl.descriptors.base import String
    tbl_mod.Table.ref = String(allow_none=False)
except Exception: pass
import openpyxl
wb=openpyxl.load_workbook(r"C:\ADHD_agent\_tmp_scripts\nav.xlsx", data_only=True, read_only=True)
out=io.open(r"C:\ADHD_agent\_tmp_scripts\xlsx.txt",'w',encoding='utf-8')
out.write("=== SHEETS ===\n")
for s in wb.sheetnames: out.write(f"  {s}\n")
# 主设计 sheet: 取第一个内容多的
sn=wb.sheetnames[0]
ws=wb[sn]
out.write(f"\n===== [{sn}] (前100非空行) =====\n")
cnt=0
for r in ws.iter_rows(values_only=True):
    cells=[str(c) for c in r if c is not None and str(c).strip()]
    if cells:
        out.write(" | ".join(cells)[:200]+"\n"); cnt+=1
    if cnt>=100: break
out.close(); print("sheets:",len(wb.sheetnames))
