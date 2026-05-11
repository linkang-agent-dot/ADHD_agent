import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

revs = [15907, 15948, 16474, 16730, 16873, 16909, 17324, 17354, 17428, 17551, 'HEAD']

def get_row_by_header(path, target_id):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['ActvOnline']
    headers = {}
    max_col = min(60, ws.max_column)
    for c in range(1, max_col+1):
        en = ws.cell(6, c).value
        if en: headers[en] = c
    for r in range(7, ws.max_row+1):
        if ws.cell(r, 1).value == target_id:
            result = {}
            for name in ['ActvType','IsOn','TimeController','CrossServerRank','NoShowRank','RankID']:
                c = headers.get(name)
                result[name] = ws.cell(r, c).value if c else '(列不存在)'
            wb.close()
            return result
    wb.close()
    return None

for target_id, label in [(101023, '黄金卷轴 大转盘'), (10071801, '蓝莲花宴 最佳酒馆')]:
    print('='*60)
    print(f'ActvID={target_id} {label}')
    print('='*60)
    prev = None
    for rev in revs:
        path = f'C:/Users/linkang/actvonline_history/AO_r{rev}.xlsx'
        try:
            info = get_row_by_header(path, target_id)
        except Exception as e:
            print(f'r{rev}: err {e}')
            continue
        if info != prev:
            print(f'\nr{rev}:')
            if info:
                for k, v in info.items():
                    oldv = prev.get(k) if isinstance(prev, dict) else None
                    mk = '*' if oldv != v else ' '
                    print(f'   {mk} {k:<18}: {v!r}')
            else:
                print('   (row not found)')
        else:
            print(f'r{rev}: (no change)')
        prev = info
    print()
