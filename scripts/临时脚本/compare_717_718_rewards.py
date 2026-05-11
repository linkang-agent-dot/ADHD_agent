import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)

# ScoreGroup details
ws = wb['ActvScoreGroup']
def collect(groups):
    result = {g: [] for g in groups}
    for r in range(7, ws.max_row+1):
        g = ws.cell(r, 2).value
        if g in result:
            result[g].append({
                'AimScore': ws.cell(r, 3).value,
                'Reward': ws.cell(r, 4).value,
            })
    return result

g717 = collect(list(range(7171, 7178)))
g718 = collect(list(range(7181, 7188)))

print('【717 vs 718 每阶段奖励档位对比】')
阶段名 = ['精英水手','城建加速','英雄晋升','研究加速','一掷千金','突发危机','声望提升']
all_same = True
for i, name in enumerate(阶段名):
    g717_id = 7171 + i
    g718_id = 7181 + i
    a, b = g717[g717_id], g718[g718_id]
    if a != b:
        all_same = False
    print(f'\nStage {i+1} {name}:  717组={g717_id} ({len(a)}档)  vs  718组={g718_id} ({len(b)}档)')
    ml = max(len(a), len(b))
    for k in range(ml):
        aa = a[k] if k < len(a) else None
        bb = b[k] if k < len(b) else None
        a_s = f'目标{aa["AimScore"]}/奖励{aa["Reward"]}' if aa else '-'
        b_s = f'目标{bb["AimScore"]}/奖励{bb["Reward"]}' if bb else '-'
        match = '✓' if aa == bb else '✗'
        print(f'    档{k+1}  717: {a_s:<25}  |  718: {b_s:<25}  {match}')

print(f'\n{"="*40}')
if all_same:
    print('🟢 717 和 718 的每阶段奖励档位完全一致')
else:
    print('🔴 717 和 718 的奖励有差异，详见上表')

# 另外也对比一下奖励 ID 785xxx 本身是否等价（检查 Reward 主表）
wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']
print('\n\n【提示】具体 RewardID 的内容需去 Reward.xlsx 查')
