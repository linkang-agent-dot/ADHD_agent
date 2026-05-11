import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
ws = wb['ActvScoreGroup']

# 收集 716 的 7161-7167 和 718 的 7181-7187 的奖励档位
def collect(groups):
    result = {g: [] for g in groups}
    for r in range(7, ws.max_row+1):
        g = ws.cell(r, 2).value
        if g in result:
            result[g].append({
                'AimScore': ws.cell(r, 3).value,
                'Reward': ws.cell(r, 4).value,
            })
    return result

g716 = collect(list(range(7161, 7168)))
g718 = collect(list(range(7181, 7188)))

print('【716 (魔女之夜本服) vs 718 (跨服2) 每阶段奖励档位对比】')
阶段名 = ['精英水手','城建加速','英雄晋升','研究加速','一掷千金','突发危机','声望提升']
for i, name in enumerate(阶段名):
    g716_id = 7161 + i
    g718_id = 7181 + i
    print(f'\nStage {i+1} {name}:  716 积分组={g716_id} ({len(g716[g716_id])}档)  vs  718 积分组={g718_id} ({len(g718[g718_id])}档)')
    max_len = max(len(g716[g716_id]), len(g718[g718_id]))
    for k in range(max_len):
        a = g716[g716_id][k] if k < len(g716[g716_id]) else None
        b = g718[g718_id][k] if k < len(g718[g718_id]) else None
        a_s = f'目标{a["AimScore"]} 奖励{a["Reward"]}' if a else '-'
        b_s = f'目标{b["AimScore"]} 奖励{b["Reward"]}' if b else '-'
        match = '✓' if a==b else '✗'
        print(f'    档{k+1}  716: {a_s:<25}  |  718: {b_s:<25}  {match}')
