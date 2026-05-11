import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 检查 ActvScore 主表 + 任何依赖 716 的地方
wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)

print('【ActvScore 主表里 ID/ContentID=716 相关行】')
for sname in ['ActvScore', 'ActvScoreGuild']:
    if sname not in wb.sheetnames: continue
    ws = wb[sname]
    print(f'\n  [{sname}]')
    for r in range(7, ws.max_row+1):
        first = ws.cell(r, 1).value
        if first == 716:
            row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
            print(f'    row{r}: {row}')

# 对照：检查 717 和 718 主表
print('\n【ActvScore 主表里 ID=717/718 的行（对照）】')
ws = wb['ActvScore']
for r in range(7, ws.max_row+1):
    v = ws.cell(r, 1).value
    if v in (716, 717, 718):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        print(f'  row{r} ID={v}: {row}')

# 看看 716 ContentID 是否还在别的子表被引用
for sname in ['ActvScoreMultiServer', 'ActvScoreGroup', 'ActvScoreTaskGroup']:
    if sname not in wb.sheetnames: continue
    ws = wb[sname]
    hit = False
    for r in range(7, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            if ws.cell(r, c).value == 716:
                if not hit:
                    print(f'\n  [{sname}] 命中 716:')
                    hit = True
                row = [ws.cell(r, cc).value for cc in range(1, ws.max_column+1)]
                print(f'    row{r}: {row}')
                break

# 看 ActvOnline 所有 Type=7 活动的 ContentID 分布
wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
print('\n【ActvOnline 所有 ActvType=7 活动】')
for r in range(7, ws_ao.max_row+1):
    t = ws_ao.cell(r, 6).value
    if t == 7:
        print(f'  ActvID={ws_ao.cell(r,1).value} 名={ws_ao.cell(r,3).value!r} IsOn={ws_ao.cell(r,7).value} ContentID={ws_ao.cell(r,5).value} 备注={ws_ao.cell(r,2).value!r}')
