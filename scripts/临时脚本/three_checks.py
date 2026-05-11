import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# === TASK 1: 尼罗-最佳酒馆 奖励道具现状 ===
print('='*70)
print('TASK 1: 尼罗-最佳酒馆 (ContentID=718) 阶段奖励道具')
print('='*70)

# Item name lookup
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
item_name = {}
for row in ws_it.iter_rows(min_row=7, max_col=3, values_only=True):
    if isinstance(row[0], int):
        item_name[row[0]] = row[1]
wb_it.close()

# 718's ScoreGroup 7181-7187 rewards → RewardID 786041-786044
wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']

# Find all reward IDs used by 718
print('\n718 每阶段奖励 (ScoreGroup 7181-7187) 引用的 RewardID:')
wb_sc = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
ws_sg = wb_sc['ActvScoreGroup']
for g in range(7181, 7188):
    print(f'\n  ScoreGroup={g}:')
    for r in range(7, ws_sg.max_row+1):
        if ws_sg.cell(r, 2).value == g:
            aim = ws_sg.cell(r, 3).value
            rid = ws_sg.cell(r, 4).value
            # lookup reward content
            items = []
            for rr in range(7, ws_rw.max_row+1):
                if ws_rw.cell(rr, 2).value == rid:
                    iid = ws_rw.cell(rr, 4).value
                    mn = ws_rw.cell(rr, 6).value
                    nm = item_name.get(iid, '?')
                    items.append(f'{iid}[{nm}]x{mn}')
            print(f'    目标{aim} → RewardID={rid} → {" + ".join(items)}')

# Check what Nile-themed items exist (search for 尼罗/埃及/猫眼/女王/法老)
print('\n\n【参考：道具表里尼罗/埃及主题道具（前30个）】')
nile_kw = ['尼罗', '埃及', '猫眼', '女王', '法老', '金字塔', '圣甲', '象形', '莲花', '沙漠']
count = 0
for iid, nm in sorted(item_name.items()):
    if nm and any(k in str(nm) for k in nile_kw):
        print(f'  {iid}: {nm}')
        count += 1
        if count >= 30: break


# === TASK 2: 大转盘里的"永生玫瑰" ===
print('\n\n' + '='*70)
print('TASK 2: 大转盘 (ContentID=1023) 里的"永生玫瑰"')
print('='*70)

# ActvLuckyWheel ContentID=1023 的 RewardGroup
wb_lw = load_workbook('C:/X3/ActvLuckyWheel.xlsx', data_only=True)
ws_lw = wb_lw['ActvLuckyWheel']
print('\n查看 ActvLuckyWheel ID=1023 的奖励池配置:')
for r in range(7, ws_lw.max_row+1):
    if ws_lw.cell(r, 1).value == 1023:
        for c in range(1, ws_lw.max_column+1):
            v = ws_lw.cell(r, c).value
            if v is not None:
                zh = ws_lw.cell(5, c).value
                en = ws_lw.cell(6, c).value
                print(f'  col{c} ({zh}/{en}): {v!r}')
        break

# Search for 永生玫瑰 in Item table
print('\n查找道具"永生玫瑰":')
for iid, nm in item_name.items():
    if nm and '永生玫瑰' in str(nm):
        print(f'  道具ID={iid} 名={nm}')

# Search ActvLuckyWheelReward for RewardGroup=319 (from 1023's config)
if 'ActvLuckyWheelReward' in wb_lw.sheetnames:
    ws_lr = wb_lw['ActvLuckyWheelReward']
    print(f'\nActvLuckyWheelReward 页签头:')
    for rr in range(5, 7):
        row = [ws_lr.cell(rr, c).value for c in range(1, min(15, ws_lr.max_column+1))]
        print(f'  row{rr}: {row}')
    print(f'\n查 Group=319 里的"永生玫瑰"相关条目:')
    for rr in range(7, ws_lr.max_row+1):
        grp = ws_lr.cell(rr, 2).value
        if grp == 319:
            iid = ws_lr.cell(rr, 3).value
            nm = item_name.get(iid, '?') if isinstance(iid, int) else '?'
            if '玫瑰' in str(nm) or '永生' in str(nm):
                row = [ws_lr.cell(rr, c).value for c in range(1, min(12, ws_lr.max_column+1))]
                print(f'  row{rr}: {row} → 道具名={nm}')


# === TASK 3: 英雄装备在D20是否开启 ===
print('\n\n' + '='*70)
print('TASK 3: 英雄装备在D20是否开启')
print('='*70)

# Search TimeCycle for 英雄装备
wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']
print('\nTimeCycle 里含"英雄装备/HeroEquip/装备"的行:')
for r in range(6, ws_tc.max_row+1):
    remark = ws_tc.cell(r, 2).value or ''
    if '装备' in str(remark) or 'equip' in str(remark).lower():
        v = ws_tc.cell(r, 1).value
        tt = ws_tc.cell(r, 5).value
        t = ws_tc.cell(r, 6).value
        dt = ws_tc.cell(r, 7).value
        d = ws_tc.cell(r, 8).value
        ison = ws_tc.cell(r, 3).value
        print(f'  TC={v} IsOn={ison} TT={tt} T={t!r} DT={dt} D={d!r} 备注={remark!r}')

# Also check FunctionUnlock / get_access for hero equipment
import os
for f in os.listdir('C:/X3'):
    fl = f.lower()
    if ('function' in fl or 'unlock' in fl or 'access' in fl) and f.endswith('.xlsx'):
        print(f'\n可能相关的功能解锁表: {f}')

# Check HeroEquip related
for f in os.listdir('C:/X3'):
    if 'HeroEquip' in f and f.endswith('.xlsx'):
        print(f'英雄装备相关表: {f}')
