import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
ws = wb['ActvScoreMulti']

print('【ActvScoreMulti — 716 vs 718 的所有字段逐项对比】')
print()
for target in (716, 718):
    print(f'--- ContentID={target} ---')
    for r in range(7, ws.max_row+1):
        if ws.cell(r, 3).value == target:
            # 全字段
            id_ = ws.cell(r, 1).value
            remark = ws.cell(r, 2).value
            stage_name = ws.cell(r, 4).value
            stage = ws.cell(r, 5).value
            duration = ws.cell(r, 6).value
            score_group = ws.cell(r, 7).value
            task_id = ws.cell(r, 8).value
            rank_id = ws.cell(r, 9).value
            print(f'  row{r} ID={id_} Stage={stage} 阶段名={stage_name!r:<10} 时长={duration} 积分组={score_group} TaskID={task_id!r:<80} RankID={rank_id} 备注={remark!r}')
    print()
