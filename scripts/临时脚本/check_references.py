import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws = wb['ActvOnline']

# 列头
headers = {}
for c in range(1, 60):
    en = ws.cell(6, c).value
    if en: headers[en] = c

# 关注的 ID
ranks_to_check = [131, 141, 142, 143, 144, 145, 134, 135, 160, 601, 602, 603, 604, 605, 606, 607]
content_ids = [717, 718]

print('='*70)
print(f'1) ActvOnline 里哪些活动的 RankID 列命中跨服/本服阶段榜')
print('='*70)
print(f'{"ActvID":>10} {"名":<14} {"Type":>4} {"IsOn":>4} {"CSR":>4} {"RankID":>7} {"ContentID":>9} {"备注":<30}')
for r in range(7, ws.max_row+1):
    aid = ws.cell(r, headers['ID']).value
    if aid is None: continue
    rank_id = ws.cell(r, headers.get('RankID', 21)).value
    if isinstance(rank_id, int) and rank_id in ranks_to_check:
        name = ws.cell(r, headers['ActvName']).value
        t = ws.cell(r, headers['ActvType']).value
        ison = ws.cell(r, headers['IsOn']).value
        csr = ws.cell(r, headers.get('CrossServerRank', 19)).value
        cid = ws.cell(r, headers['ContentID']).value
        remark = ws.cell(r, 2).value
        print(f'{aid:>10} {str(name)[:14]:<14} {t!s:>4} {ison!s:>4} {csr!s:>4} {rank_id:>7} {cid!s:>9} {str(remark)[:30]:<30}')

print()
print('='*70)
print(f'2) ActvOnline 里哪些活动 ContentID=717 或 718')
print('='*70)
print(f'{"ActvID":>10} {"名":<14} {"Type":>4} {"IsOn":>4} {"CSR":>4} {"RankID":>7} {"ContentID":>9}')
for r in range(7, ws.max_row+1):
    aid = ws.cell(r, headers['ID']).value
    if aid is None: continue
    cid = ws.cell(r, headers['ContentID']).value
    if cid in content_ids:
        name = ws.cell(r, headers['ActvName']).value
        t = ws.cell(r, headers['ActvType']).value
        ison = ws.cell(r, headers['IsOn']).value
        csr = ws.cell(r, headers.get('CrossServerRank', 19)).value
        rank_id = ws.cell(r, headers.get('RankID', 21)).value
        print(f'{aid:>10} {str(name)[:14]:<14} {t!s:>4} {ison!s:>4} {csr!s:>4} {rank_id!s:>7} {cid!s:>9}')

print()
print('='*70)
print(f'3) 扫全表, 所有列里出现 131/141-145/134/135/601-607 的位置')
print('='*70)
tgt = set(ranks_to_check)
hits = {}
for r in range(7, ws.max_row+1):
    for c in range(1, 60):
        v = ws.cell(r, c).value
        if v in tgt:
            aid = ws.cell(r, 1).value
            colname = ws.cell(6, c).value or f'col{c}'
            hits.setdefault(v, []).append((aid, c, colname))

for rank_id in sorted(tgt):
    locs = hits.get(rank_id, [])
    if locs:
        print(f'\n  RankID={rank_id} ({len(locs)} 处命中):')
        for aid, c, cn in locs[:15]:
            print(f'    ActvID={aid} col{c}({cn})')
