import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

def scan(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['ActvOnline']
    headers = {}
    for c in range(1, 60):
        en = ws.cell(6, c).value
        if en: headers[en] = c
    rows = {}
    for r in range(7, ws.max_row+1):
        aid = ws.cell(r, headers['ID']).value
        if aid is None: continue
        rows[aid] = {
            'ActvName': ws.cell(r, headers.get('ActvName',3)).value,
            'ActvType': ws.cell(r, headers['ActvType']).value,
            'IsOn': ws.cell(r, headers['IsOn']).value,
            'TimeController': ws.cell(r, headers['TimeController']).value,
            'CrossServerRank': ws.cell(r, headers.get('CrossServerRank', 19)).value,
            'NoShowRank': ws.cell(r, headers.get('NoShowRank', 20)).value,
            'RankID': ws.cell(r, headers.get('RankID', 21)).value,
        }
    wb.close()
    return rows

r17552 = scan('C:/Users/linkang/actvonline_history/AO_r17552.xlsx')
r17553 = scan('C:/Users/linkang/actvonline_history/AO_r17553.xlsx')
r17554 = scan('C:/Users/linkang/actvonline_history/AO_r17554.xlsx')

print('=== r17552 → r17553 差异 ===')
for aid in sorted(set(r17552) | set(r17553)):
    o = r17552.get(aid) or {}
    n = r17553.get(aid) or {}
    for k in ['ActvType','IsOn','TimeController','CrossServerRank','NoShowRank','RankID']:
        if o.get(k) != n.get(k):
            print(f'  {aid} {n.get("ActvName") or o.get("ActvName")!r} {k}: {o.get(k)!r} -> {n.get(k)!r}')

print('\n=== r17553 → r17554 差异 ===')
for aid in sorted(set(r17553) | set(r17554)):
    o = r17553.get(aid) or {}
    n = r17554.get(aid) or {}
    for k in ['ActvType','IsOn','TimeController','CrossServerRank','NoShowRank','RankID']:
        if o.get(k) != n.get(k):
            print(f'  {aid} {n.get("ActvName") or o.get("ActvName")!r} {k}: {o.get(k)!r} -> {n.get(k)!r}')

print('\n=== r17554 (当前) 尼罗系列 10 个活动状态 ===')
nile_ids = [101023, 101024, 102234, 105602, 101825, 105010, 10071801, 101334, 101335, 100594]
for aid in nile_ids:
    r = r17554.get(aid)
    if r:
        print(f'  {aid:>9} {r["ActvName"]!r:<15} Type={r["ActvType"]:<3} IsOn={r["IsOn"]:<3} TC={r["TimeController"]!s:<6} CSR={r["CrossServerRank"]!r:<5} RankID={r["RankID"]!r}')
