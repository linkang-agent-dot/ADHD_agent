import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

print('加载 Reward.xlsx...')
wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']

targets = set(range(785301, 785309))
rewards = {tid: [] for tid in targets}
for r in range(7, ws_rw.max_row+1):
    rid = ws_rw.cell(r, 2).value
    if rid in targets:
        rewards[rid].append({
            'ItemType': ws_rw.cell(r, 3).value,
            'ItemID': ws_rw.cell(r, 4).value,
            'MinNum': ws_rw.cell(r, 6).value,
            'MaxNum': ws_rw.cell(r, 7).value,
            'DropType': ws_rw.cell(r, 8).value,
            'DropPara': ws_rw.cell(r, 9).value,
        })
print('Reward.xlsx done')

print('加载 Item.xlsx...')
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item'] if 'Item' in wb_it.sheetnames else wb_it[wb_it.sheetnames[0]]
# 列头
name_col = None
id_col = 1
for c in range(1, 30):
    zh = ws_it.cell(5, c).value
    en = ws_it.cell(6, c).value
    if zh and '名' in str(zh):
        name_col = c
        break
    if en and ('Name' in str(en) or 'TXT_Name' in str(en)):
        name_col = c
        break
print(f'Item.xlsx 列: id_col={id_col}, name_col={name_col}')

item_name = {}
row_count = 0
for row in ws_it.iter_rows(min_row=7, max_col=max(id_col, name_col or 1)+1, values_only=True):
    row_count += 1
    iid = row[id_col-1]
    if isinstance(iid, int) and name_col:
        item_name[iid] = row[name_col-1]

wb_it.close()
print(f'Item 读取 {row_count} 行，收集到 {len(item_name)} 个道具名')

排名档位 = [(785301, '第1名'), (785302, '第2名'), (785303, '第3名'),
         (785304, '第4-5名'), (785305, '第6-10名'), (785306, '第11-20名'),
         (785307, '第21-50名'), (785308, '第51-100名')]

print('\n========================================')
print('RankID=160 通用本服节庆酒馆 排名奖励')
print('========================================')
for rid, label in 排名档位:
    items = rewards.get(rid, [])
    print(f'\n{label} (RewardID={rid}): {len(items)}项')
    for it in items:
        name = item_name.get(it['ItemID'], '?')
        n = f"{it['MinNum']}" if it['MaxNum'] in (None, it['MinNum']) else f"{it['MinNum']}~{it['MaxNum']}"
        d = '固定' if it['DropType']==1 else f"权重{it['DropPara']}"
        itype = {1:'道具', 2:'蓝图', 3:'子包'}.get(it['ItemType'], str(it['ItemType']))
        print(f"  {itype} ID={it['ItemID']} {name!r} x{n} ({d})")
