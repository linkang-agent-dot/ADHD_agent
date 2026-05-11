import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/Rank.xlsx', data_only=True)
ws = wb['RankCfg']
print('【本服/阶段榜相关 ID 定义】')
for r in range(6, ws.max_row+1):
    v = ws.cell(r, 1).value
    if isinstance(v, int) and ((130 <= v <= 160) or (601 <= v <= 620)):
        rt = ws.cell(r, 4).value
        cat = '6本服' if rt == 6 else ('12跨服' if rt == 12 else f'{rt}')
        print(f'  ID={v} RankType={cat:<7} 备注={ws.cell(r,2).value!r}')
