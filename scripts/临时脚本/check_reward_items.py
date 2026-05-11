import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']

targets = set(range(786031, 786035)) | set(range(786041, 786045))
rewards = {tid: [] for tid in targets}
for r in range(7, ws_rw.max_row+1):
    rid = ws_rw.cell(r, 2).value
    if rid in targets:
        rewards[rid].append({
            'ItemType': ws_rw.cell(r, 3).value,
            'ItemID': ws_rw.cell(r, 4).value,
            'MinNum': ws_rw.cell(r, 6).value,
            'MaxNum': ws_rw.cell(r, 7).value,
            'DropType': ws_rw.cell(r, 8).value,
            'DropPara': ws_rw.cell(r, 9).value,
        })

# Load Item names
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
item_name = {}
for row in ws_it.iter_rows(min_row=7, max_col=3, values_only=True):
    iid = row[0]
    if isinstance(iid, int):
        item_name[iid] = row[1]
wb_it.close()

def fmt(items):
    parts = []
    for it in items:
        nm = item_name.get(it['ItemID'], '?')
        n = f"{it['MinNum']}" if it['MaxNum'] in (None, it['MinNum']) else f"{it['MinNum']}~{it['MaxNum']}"
        parts.append(f"{it['ItemID']}[{nm}]x{n}")
    return ' + '.join(parts)

print('【717 档位奖励】')
for rid in sorted(range(786031, 786035)):
    print(f'  {rid}: {fmt(rewards.get(rid, []))}')

print('\n【718 档位奖励】')
for rid in sorted(range(786041, 786045)):
    print(f'  {rid}: {fmt(rewards.get(rid, []))}')

print('\n【档位对应对比 (717档位N vs 718档位N 内容是否一致)】')
for k in range(4):
    a = rewards.get(786031+k, [])
    b = rewards.get(786041+k, [])
    # 比对 (ItemID, Count)
    a_set = sorted((i['ItemID'], i['MinNum'], i['MaxNum']) for i in a)
    b_set = sorted((i['ItemID'], i['MinNum'], i['MaxNum']) for i in b)
    match = '✓ 一致' if a_set == b_set else '✗ 不同'
    print(f'  档{k+1}: {match}')
    if a_set != b_set:
        print(f'    717(R{786031+k}): {fmt(a)}')
        print(f'    718(R{786041+k}): {fmt(b)}')
