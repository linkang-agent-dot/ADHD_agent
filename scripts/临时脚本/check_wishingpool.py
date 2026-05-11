import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

revs = [15907, 15948, 16474, 16730, 16873, 16909, 17354, 17428, 17551, 'HEAD']

def scan_actvtype_50(path):
    """Return list of (ActvID, ActvName, ActvType, IsOn, CrossServerRank, RankID, TC)"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['ActvOnline']
    # headers
    headers = {}
    for c in range(1, 60):
        en = ws.cell(6, c).value
        if en: headers[en] = c
    id_c = headers.get('ID', 1)
    name_c = headers.get('ActvName', 3)
    type_c = headers.get('ActvType', 6)
    ison_c = headers.get('IsOn', 7)
    tc_c = headers.get('TimeController', 8)
    crs_c = headers.get('CrossServerRank')
    rank_c = headers.get('RankID')

    rows = []
    for r in range(7, ws.max_row+1):
        t = ws.cell(r, type_c).value
        if t == 50:
            rows.append({
                'ID': ws.cell(r, id_c).value,
                'ActvName': ws.cell(r, name_c).value,
                'IsOn': ws.cell(r, ison_c).value,
                'TimeController': ws.cell(r, tc_c).value,
                'CrossServerRank': ws.cell(r, crs_c).value if crs_c else '?',
                'RankID': ws.cell(r, rank_c).value if rank_c else '?',
            })
    wb.close()
    return rows

print('='*80)
print('ActvType=50 (许愿池抽奖活动) 在各历史版本里的所有活动及跨服配置')
print('='*80)

for rev in revs:
    path = f'C:/Users/linkang/actvonline_history/AO_r{rev}.xlsx'
    try:
        rows = scan_actvtype_50(path)
    except Exception as e:
        print(f'\nr{rev}: 加载失败 {e}')
        continue
    print(f'\nr{rev}: 共 {len(rows)} 个 Type=50 活动')
    for r in rows:
        print(f"  ID={r['ID']} {str(r['ActvName'])[:10]:<10} IsOn={r['IsOn']!s:<5} TC={r['TimeController']!s:<6} CrossServerRank={r['CrossServerRank']!r:<5} RankID={r['RankID']!r}")
