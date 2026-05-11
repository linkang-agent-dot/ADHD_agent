import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 对比 r17551 (我改好的) vs r17552 (你最新提交) 看什么变了
def scan(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['ActvOnline']
    headers = {}
    for c in range(1, 60):
        en = ws.cell(6, c).value
        if en: headers[en] = c
    rows = {}
    for r in range(7, ws.max_row+1):
        aid = ws.cell(r, headers['ID']).value
        if aid is None: continue
        rows[aid] = {
            'row': r,
            'ActvName': ws.cell(r, headers.get('ActvName',3)).value,
            'ActvType': ws.cell(r, headers['ActvType']).value,
            'IsOn': ws.cell(r, headers['IsOn']).value,
            'TimeController': ws.cell(r, headers['TimeController']).value,
            'CrossServerRank': ws.cell(r, headers.get('CrossServerRank', 19)).value,
            'NoShowRank': ws.cell(r, headers.get('NoShowRank', 20)).value,
            'RankID': ws.cell(r, headers.get('RankID', 21)).value,
        }
    wb.close()
    return rows

old = scan('C:/Users/linkang/actvonline_history/AO_r17551.xlsx')
new = scan('C:/Users/linkang/actvonline_history/AO_r17552.xlsx')

# 对比
print('='*80)
print('r17551 (我改好的还原版) → r17552 (你 X3NEW-123 修复提交) 差异')
print('='*80)
changed = []
for aid in sorted(set(old) | set(new)):
    o = old.get(aid)
    n = new.get(aid)
    if o != n:
        changed.append(aid)
print(f'\n共 {len(changed)} 个活动被改动:\n')
for aid in changed:
    o = old.get(aid) or {}
    n = new.get(aid) or {}
    name = n.get('ActvName') or o.get('ActvName')
    print(f'ActvID={aid} {name!r}')
    for k in ['ActvType','IsOn','TimeController','CrossServerRank','NoShowRank','RankID']:
        ov = o.get(k)
        nv = n.get(k)
        if ov != nv:
            print(f'    {k}: {ov!r} -> {nv!r}')
    print()

# 特别看 105010
print('\n【105010 许愿池 当前 r17552 状态】')
if 105010 in new:
    for k, v in new[105010].items():
        print(f'  {k}: {v!r}')
