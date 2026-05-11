import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# Item 表查"传奇英雄装备宝箱"
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
item_name = {}
equip_items = []
for row in ws_it.iter_rows(min_row=7, max_col=30, values_only=True):
    iid = row[0]
    nm = row[1] if len(row) > 1 else None
    if isinstance(iid, int):
        item_name[iid] = nm
    if nm and ('英雄装备' in str(nm) or '装备宝箱' in str(nm)):
        equip_items.append((iid, nm, row[3] if len(row) > 3 else None, row[4] if len(row) > 4 else None))
wb_it.close()

print('【道具表里含"英雄装备"或"装备宝箱"的道具】')
for iid, nm, typ, desc in equip_items:
    print(f'  {iid}: {nm}')

# 具体查"传奇英雄装备宝箱"
print('\n【精确搜索"传奇英雄装备宝箱"】')
wb_it2 = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it2 = wb_it2['Item']
for r in range(7, ws_it2.max_row+1):
    nm = ws_it2.cell(r, 2).value
    if nm and '传奇英雄装备宝箱' in str(nm):
        print(f'  道具ID={ws_it2.cell(r,1).value} 名={nm}')
        for c in range(1, 30):
            v = ws_it2.cell(r, c).value
            if v is not None:
                zh = ws_it2.cell(5, c).value
                en = ws_it2.cell(6, c).value
                print(f'    col{c} ({zh}/{en}): {v!r}')
        print()
wb_it2.close()

# 在尼罗相关表里搜索这个道具
# 1) 大转盘 Group=319
print('\n【大转盘 Group=319 里是否有英雄装备宝箱】')
wb_lw = load_workbook('C:/X3/ActvLuckyWheel.xlsx', data_only=True)
if 'ActvLuckyWheelReward' in wb_lw.sheetnames:
    ws_lr = wb_lw['ActvLuckyWheelReward']
    for r in range(7, ws_lr.max_row+1):
        if ws_lr.cell(r, 2).value == 319:
            iid = ws_lr.cell(r, 3).value
            nm = item_name.get(iid, '?')
            if '装备' in str(nm):
                row = [ws_lr.cell(r, c).value for c in range(1, 9)]
                print(f'  row{r}: {row} → {nm}')

# 2) 兑换商店 ContentID=1334 里有没有
print('\n【兑换商店 ContentID=1334 里是否有英雄装备宝箱】')
wb_ex = load_workbook('C:/X3/ActvExchange.xlsx', data_only=True)
ws_ex = wb_ex['ActvExchange']
for r in range(7, ws_ex.max_row+1):
    cid = ws_ex.cell(r, 2).value
    if cid == 1334:
        iid = ws_ex.cell(r, 4).value
        nm = item_name.get(iid, '?')
        if '装备' in str(nm):
            print(f'  row{r} ContentID=1334 道具={iid} {nm} 数量={ws_ex.cell(r,6).value} 消耗={ws_ex.cell(r,8).value}猫眼石币')

# 3) 英雄装备功能解锁时间
print('\n【FunctionUnlock 里"英雄装备/HeroEquip"相关】')
wb_fu = load_workbook('C:/X3/FunctionUnlock.xlsx', data_only=True)
ws_fu = wb_fu[wb_fu.sheetnames[0]]
for r in range(7, ws_fu.max_row+1):
    row_text = ' '.join(str(ws_fu.cell(r, c).value or '') for c in range(1, min(ws_fu.max_column+1, 15)))
    if any(k in row_text.lower() for k in ['英雄装备', 'heroequip', 'hero_equip']):
        row = [ws_fu.cell(r, c).value for c in range(1, min(ws_fu.max_column+1, 15))]
        print(f'  row{r}: {row}')

# search HeroEquip.xlsx for unlock condition
print('\n【HeroEquip.xlsx 结构】')
wb_he = load_workbook('C:/X3/HeroEquip.xlsx', data_only=True, read_only=True)
print(f'  Sheets: {wb_he.sheetnames}')
for sname in wb_he.sheetnames[:3]:
    ws_he = wb_he[sname]
    print(f'\n  [{sname}] row5: {[ws_he.cell(5,c).value for c in range(1, min(ws_he.max_column+1, 12))]}')
    print(f'  [{sname}] row6: {[ws_he.cell(6,c).value for c in range(1, min(ws_he.max_column+1, 12))]}')
wb_he.close()
