import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# Item name lookup
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
item_name = {}
for row in ws_it.iter_rows(min_row=7, max_col=5, values_only=True):
    if isinstance(row[0], int):
        item_name[row[0]] = row[1]
wb_it.close()

# 1) ActvLuckyWheel ID=1023 的 ServerRank 字段
print('='*60)
print('1) ActvLuckyWheel ID=1023 配置')
print('='*60)
wb_lw = load_workbook('C:/X3/ActvLuckyWheel.xlsx', data_only=True)
ws_lw = wb_lw['ActvLuckyWheel']
for r in range(7, ws_lw.max_row+1):
    if ws_lw.cell(r, 1).value == 1023:
        for c in range(1, ws_lw.max_column+1):
            v = ws_lw.cell(r, c).value
            if v is not None:
                zh = ws_lw.cell(5, c).value
                en = ws_lw.cell(6, c).value
                print(f'  col{c} ({zh}/{en}): {v!r}')
        break

# 2) ServerRank=169 的排行榜奖励（本服排行）
print(f'\n{"="*60}')
print('2) RankID=169 本服排行榜奖励')
print('='*60)
wb_rk = load_workbook('C:/X3/Rank.xlsx', data_only=True)
ws_rs = wb_rk['RankRewardSlotCfg']
print(f'{"ID":>8} {"起":>3} {"止":>3} {"RewardID":>10}')
reward_ids_169 = []
for r in range(7, ws_rs.max_row+1):
    if ws_rs.cell(r, 2).value == 169:
        rid = ws_rs.cell(r, 5).value
        reward_ids_169.append(rid)
        print(f'{ws_rs.cell(r,1).value!s:>8} {ws_rs.cell(r,3).value!s:>3} {ws_rs.cell(r,4).value!s:>3} {rid!s:>10}')

# 3) 这些奖励的具体内容 — 找皮肤道具
print(f'\n{"="*60}')
print('3) RankID=169 排行榜奖励里的皮肤/英雄道具')
print('='*60)
wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']

for rid in reward_ids_169:
    items = []
    for r in range(7, ws_rw.max_row+1):
        if ws_rw.cell(r, 2).value == rid:
            iid = ws_rw.cell(r, 4).value
            mn = ws_rw.cell(r, 6).value
            nm = item_name.get(iid, '?')
            items.append((iid, nm, mn))
    skin_items = [(iid, nm, mn) for iid, nm, mn in items if any(k in str(nm) for k in ['皮肤', '永久', '纪念卡', '时装', '外观'])]
    if skin_items:
        print(f'\n  RewardID={rid}:')
        for iid, nm, mn in items:
            marker = ' ★皮肤' if any(k in str(nm) for k in ['皮肤', '永久', '纪念卡', '时装', '外观']) else ''
            print(f'    {iid} {nm} x{mn}{marker}')

# 4) 更广泛：大转盘 Group=319 奖池 + OtherRewardGroup=3019 额外奖励里有没有皮肤
print(f'\n{"="*60}')
print('4) 大转盘 Group=319 完整奖池 + OtherRewardGroup=3019 额外奖励')
print('='*60)

# 额外奖励组
if 'ActvLuckyWheelOtherReward' in wb_lw.sheetnames:
    ws_or = wb_lw['ActvLuckyWheelOtherReward']
    print(f'\nOtherReward 列头:')
    print(f'  row5: {[ws_or.cell(5,c).value for c in range(1,10)]}')
    print(f'  row6: {[ws_or.cell(6,c).value for c in range(1,10)]}')
    print(f'\nOtherRewardGroup=3019:')
    for r in range(7, ws_or.max_row+1):
        if ws_or.cell(r, 2).value == 3019:
            row = [ws_or.cell(r, c).value for c in range(1, 10)]
            iid = ws_or.cell(r, 3).value
            nm = item_name.get(iid, '?') if isinstance(iid, int) else '?'
            print(f'  row{r}: {row} → {nm}')

# 5) 皮肤对应的英雄 — 搜索道具表中所有 53xxx 段的皮肤
print(f'\n{"="*60}')
print('5) 道具表里 53xxx / 530xx 段皮肤道具（与英雄关联）')
print('='*60)
wb_it3 = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it3 = wb_it3['Item']
# 搜索尼罗相关皮肤
nile_skin_kw = ['尼罗', '埃及', '法老', '女王', '莲花', '金字塔', '圣甲', '沙漠']
print('\n尼罗主题皮肤:')
for r in range(7, ws_it3.max_row+1):
    iid = ws_it3.cell(r, 1).value
    nm = ws_it3.cell(r, 2).value
    if nm and any(k in str(nm) for k in nile_skin_kw) and any(k in str(nm) for k in ['皮肤', '永久', '（永久）', '纪念卡']):
        desc = ws_it3.cell(r, 4).value
        print(f'  {iid}: {nm} → {desc!r}')
wb_it3.close()
