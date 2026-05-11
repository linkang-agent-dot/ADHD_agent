import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 1) ConstConfig 搜索英雄装备
print('='*60)
print('1) ConstConfig 搜索英雄装备/HeroEquip')
print('='*60)
for f in os.listdir('C:/X3'):
    if 'Const' in f and f.endswith('.xlsx') and not f.startswith('~'):
        print(f'\n文件: {f}')
        wb = load_workbook(f'C:/X3/{f}', data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, ws.max_row+1):
                row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, min(ws.max_column+1, 10)))
                if any(k in row_text.lower() for k in ['heroequip', 'hero_equip', '英雄装备', '装备解锁', 'equip_unlock']):
                    row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 10))]
                    print(f'  [{sname}] row{r}: {row}')
        wb.close()

# 2) Building.xlsx 搜索英雄装备解锁
print('\n' + '='*60)
print('2) Building.xlsx 搜索英雄装备相关')
print('='*60)
if os.path.exists('C:/X3/Building.xlsx'):
    wb = load_workbook('C:/X3/Building.xlsx', data_only=True, read_only=True)
    print(f'Sheets: {wb.sheetnames}')
    ws = wb[wb.sheetnames[0]]
    print(f'列头 row5: {[ws.cell(5,c).value for c in range(1, min(ws.max_column+1, 15))]}')
    print(f'列头 row6: {[ws.cell(6,c).value for c in range(1, min(ws.max_column+1, 15))]}')
    for r in range(7, ws.max_row+1):
        row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, min(ws.max_column+1, 15)))
        if any(k in row_text.lower() for k in ['heroequip', '英雄装备', 'hero_equip']):
            row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 15))]
            print(f'  row{r}: {row}')
    wb.close()

# 3) FunctionUnlock 完整搜索 — 用更宽泛的关键词
print('\n' + '='*60)
print('3) FunctionUnlock 更宽泛搜索 (装备/equip/166004)')
print('='*60)
wb = load_workbook('C:/X3/FunctionUnlock.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
for r in range(7, ws.max_row+1):
    row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, min(ws.max_column+1, 15)))
    if any(k in row_text.lower() for k in ['装备', 'equip', '166004', '166006']):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 15))]
        print(f'  row{r}: {row}')

# 4) HeroEquip.xlsx — 看有没有解锁等级/条件字段
print('\n' + '='*60)
print('4) HeroEquipSlot 页签 (可能有解锁条件)')
print('='*60)
wb_he = load_workbook('C:/X3/HeroEquip.xlsx', data_only=True, read_only=True)
if 'HeroEquipSlot' in wb_he.sheetnames:
    ws = wb_he['HeroEquipSlot']
    print(f'列头 row4: {[ws.cell(4,c).value for c in range(1, min(ws.max_column+1, 12))]}')
    print(f'列头 row5: {[ws.cell(5,c).value for c in range(1, min(ws.max_column+1, 12))]}')
    print(f'列头 row6: {[ws.cell(6,c).value for c in range(1, min(ws.max_column+1, 12))]}')
    for r in range(7, min(ws.max_row+1, 20)):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 12))]
        print(f'  row{r}: {row}')

# 5) get_access_group 搜索
print('\n' + '='*60)
print('5) get_access_group 搜索装备相关')
print('='*60)
for f in os.listdir('C:/X3'):
    if 'access' in f.lower() and f.endswith('.xlsx') and not f.startswith('~'):
        print(f'\n文件: {f}')
        wb2 = load_workbook(f'C:/X3/{f}', data_only=True, read_only=True)
        for sname in wb2.sheetnames[:2]:
            ws2 = wb2[sname]
            for r in range(1, ws2.max_row+1):
                row_text = ' '.join(str(ws2.cell(r, c).value or '') for c in range(1, min(ws2.max_column+1, 10)))
                if any(k in row_text.lower() for k in ['装备', 'equip', 'heroequip']):
                    row = [ws2.cell(r, c).value for c in range(1, min(ws2.max_column+1, 10))]
                    print(f'  [{sname}] row{r}: {row}')
        wb2.close()
wb_he.close()
