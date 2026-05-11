import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 1) ActvOnline 里永恒之岛的完整信息
wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws_ao.cell(6, c).value
    if en: headers[en] = c

print('='*80)
print('永恒之岛（奇观赛季）活动信息')
print('='*80)
for r in range(7, ws_ao.max_row+1):
    aid = ws_ao.cell(r, 1).value
    remark = str(ws_ao.cell(r, 2).value or '')
    if '永恒之岛' in str(ws_ao.cell(r, 3).value or '') or '奇观' in remark:
        print(f'\nActvID={aid} 名={ws_ao.cell(r,3).value!r} Type={ws_ao.cell(r,6).value} IsOn={ws_ao.cell(r,7).value}')
        print(f'  备注: {remark}')
        print(f'  ContentID={ws_ao.cell(r,5).value} TC={ws_ao.cell(r,8).value}')
        tc_id = ws_ao.cell(r, 8).value
        # 其他关键字段
        for col_name in ['CrossServerRank','RankID','GroupId']:
            c = headers.get(col_name)
            if c:
                v = ws_ao.cell(r, c).value
                if v is not None:
                    print(f'  {col_name}={v}')

# 2) 搜索相关的配置表
import os
print(f'\n{"="*80}')
print('搜索奇观/永恒之岛相关配置表')
print(f'{"="*80}')
for f in os.listdir('C:/X3'):
    fl = f.lower()
    if ('invasion' in fl or 'wonder' in fl or 'nest' in fl or 'island' in fl) and f.endswith('.xlsx') and not f.startswith('~'):
        print(f'  {f}')

# 3) ActvScore 里 ContentID=1207/1208 的积分配置
wb_sc = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
print(f'\n{"="*80}')
print('ActvScore 里奇观赛季的积分配置')
print(f'{"="*80}')
for sname in ['ActvScore', 'ActvScoreGuild']:
    ws = wb_sc[sname]
    for r in range(7, ws.max_row+1):
        v = ws.cell(r, 1).value
        if v in (1207, 1208):
            row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 8))]
            print(f'  [{sname}] ID={v}: {row}')

# 4) 看看奇观赛季的 Pack/礼包引用
# search ActvOnline 里 ContentID=1207/1208 关联的所有字段
print(f'\n{"="*80}')
print('奇观赛季关联的礼包/链式礼包')
print(f'{"="*80}')
for r in range(7, ws_ao.max_row+1):
    cid = ws_ao.cell(r, headers['ContentID']).value
    if cid in (1207, 1208):
        aid = ws_ao.cell(r, 1).value
        # 扫所有非空列
        for c in range(1, 50):
            v = ws_ao.cell(r, c).value
            zh = ws_ao.cell(5, c).value
            en = ws_ao.cell(6, c).value
            if v is not None and any(k in str(zh or '').lower() + str(en or '').lower() for k in ['pack','礼包','链式','chain','累充','recharge']):
                print(f'  ActvID={aid} col{c}({zh}/{en})={v}')

# 5) 净化试炼(修女挑战) 信息
print(f'\n{"="*80}')
print('净化试炼（修女挑战）活动信息')
print(f'{"="*80}')
for r in range(7, ws_ao.max_row+1):
    aid = ws_ao.cell(r, 1).value
    if aid == 100808:
        for c in range(1, 50):
            v = ws_ao.cell(r, c).value
            if v is not None:
                zh = ws_ao.cell(5, c).value
                en = ws_ao.cell(6, c).value
                print(f'  col{c}({zh}/{en})={v!r}')

# 6) 武道冠军之路
print(f'\n{"="*80}')
print('武道冠军之路 活动信息')
print(f'{"="*80}')
for r in range(7, ws_ao.max_row+1):
    aid = ws_ao.cell(r, 1).value
    remark = str(ws_ao.cell(r, 2).value or '')
    if '武道冠军' in str(ws_ao.cell(r, 3).value or '') or '武道冠军' in remark:
        for c in range(1, 50):
            v = ws_ao.cell(r, c).value
            if v is not None:
                zh = ws_ao.cell(5, c).value
                en = ws_ao.cell(6, c).value
                print(f'  col{c}({zh}/{en})={v!r}')
