import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 打印 ActvScoreMulti 全部数据，按 ContentID 分组
wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
ws = wb['ActvScoreMulti']
print('【ActvScoreMulti 全表 (ContentID, 备注, Stage, RankID)】')
print(f'{"row":>4} {"ID":>6} {"ContentID":>10} {"Stage":>5} {"RankID":>6} {"备注":<30}')
for r in range(7, ws.max_row+1):
    rid = ws.cell(r, 1).value
    remark = ws.cell(r, 2).value
    cid = ws.cell(r, 3).value
    stage = ws.cell(r, 5).value
    rankid = ws.cell(r, 9).value
    if cid is not None:
        print(f'{r:>4} {rid!s:>6} {cid!s:>10} {stage!s:>5} {rankid!s:>6} {str(remark or ""):<30}')

# 也看一下 RankCfg 里 601~607 的定义，确认是跨服
wb_rk = load_workbook('C:/X3/Rank.xlsx', data_only=True)
ws_rk = wb_rk['RankCfg']
print('\n\n【RankCfg ID 601~609 定义】')
for r in range(6, ws_rk.max_row+1):
    v = ws_rk.cell(r, 1).value
    if isinstance(v, int) and 600 <= v <= 620:
        rt = ws_rk.cell(r, 4).value
        cat = '6本服' if rt == 6 else ('12跨服' if rt == 12 else f'{rt}')
        print(f'  ID={v} RankType={cat} 备注={ws_rk.cell(r,2).value!r} MaxSize={ws_rk.cell(r,7).value}')
