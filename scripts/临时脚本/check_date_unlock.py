import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 1) FunctionUnlock.xlsx 搜索约会相关
wb = load_workbook('C:/X3/FunctionUnlock.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print(f'\n{wb.sheetnames[0]} 列头:')
for r in range(1, 7):
    row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 15))]
    print(f'  row{r}: {row}')

print(f'\n搜索约会/Date/玫瑰/romance相关:')
for r in range(7, ws.max_row+1):
    row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, min(ws.max_column+1, 15)))
    if any(k in row_text.lower() for k in ['约会', 'date', '玫瑰', 'romance', '1022']):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 15))]
        print(f'  row{r}: {row}')

# 2) Date.xlsx 搜索
import os
for f in os.listdir('C:/X3'):
    if ('Date' in f or 'date' in f.lower() or 'Romance' in f or 'romance' in f.lower()) and f.endswith('.xlsx') and not f.startswith('~'):
        print(f'\n找到相关表: {f}')
        try:
            wb2 = load_workbook(f'C:/X3/{f}', data_only=True, read_only=True)
            print(f'  Sheets: {wb2.sheetnames}')
            ws2 = wb2[wb2.sheetnames[0]]
            print(f'  row5: {[ws2.cell(5,c).value for c in range(1, min(ws2.max_column+1, 12))]}')
            print(f'  row6: {[ws2.cell(6,c).value for c in range(1, min(ws2.max_column+1, 12))]}')
            wb2.close()
        except Exception as e:
            print(f'  打开失败: {e}')

# 3) TimeCycle 搜索约会
wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']
print(f'\nTimeCycle 里含约会/Date/romance的行:')
for r in range(6, ws_tc.max_row+1):
    remark = str(ws_tc.cell(r, 2).value or '')
    if any(k in remark.lower() for k in ['约会', 'date', 'romance']):
        v = ws_tc.cell(r, 1).value
        tt = ws_tc.cell(r, 5).value
        t = ws_tc.cell(r, 6).value
        ison = ws_tc.cell(r, 3).value
        print(f'  TC={v} IsOn={ison} TT={tt} T={t!r} 备注={remark!r}')

# 4) ActvOnline 搜索约会
wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws_ao.cell(6, c).value
    if en: headers[en] = c
print(f'\nActvOnline 里含约会/Date/romance的活动:')
for r in range(7, ws_ao.max_row+1):
    name = str(ws_ao.cell(r, headers.get('ActvName', 3)).value or '')
    remark = str(ws_ao.cell(r, 2).value or '')
    combo = name + remark
    if any(k in combo.lower() for k in ['约会', 'date', 'romance', '玫瑰']):
        aid = ws_ao.cell(r, 1).value
        ison = ws_ao.cell(r, headers['IsOn']).value
        tc = ws_ao.cell(r, headers['TimeController']).value
        t = ws_ao.cell(r, headers['ActvType']).value
        print(f'  ActvID={aid} 名={name!r} Type={t} IsOn={ison} TC={tc} 备注={remark!r}')
