import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 1) Item 表里 1022 永生玫瑰的完整信息
wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
# 列头
headers_it = {}
for c in range(1, 30):
    zh = ws_it.cell(5, c).value
    en = ws_it.cell(6, c).value
    if en: headers_it[en] = c
    if zh: headers_it[zh] = c

print('【Item.xlsx 道具 1022 永生玫瑰 完整信息】')
for r in range(7, ws_it.max_row+1):
    if ws_it.cell(r, 1).value == 1022:
        for c in range(1, 30):
            v = ws_it.cell(r, c).value
            if v is not None:
                zh = ws_it.cell(5, c).value
                en = ws_it.cell(6, c).value
                print(f'  col{c} ({zh}/{en}): {v!r}')
        break
wb_it.close()

# 2) 在哪些奖励/掉落里出现 1022
print('\n【Reward.xlsx 里 ItemID=1022 的所有条目】')
wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']
for r in range(7, ws_rw.max_row+1):
    iid = ws_rw.cell(r, 4).value
    if iid == 1022:
        rid = ws_rw.cell(r, 2).value
        mn = ws_rw.cell(r, 6).value
        mx = ws_rw.cell(r, 7).value
        dt = ws_rw.cell(r, 8).value
        dp = ws_rw.cell(r, 9).value
        print(f'  RewardID={rid} 数量={mn}~{mx} DropType={dt} DropPara={dp}')

# 3) 在 ActvLuckyWheelReward 里的出现
print('\n【ActvLuckyWheelReward 里 ItemID=1022 的所有条目】')
wb_lw = load_workbook('C:/X3/ActvLuckyWheel.xlsx', data_only=True)
if 'ActvLuckyWheelReward' in wb_lw.sheetnames:
    ws_lr = wb_lw['ActvLuckyWheelReward']
    for r in range(7, ws_lr.max_row+1):
        if ws_lr.cell(r, 3).value == 1022:
            row = [ws_lr.cell(r, c).value for c in range(1, 9)]
            grp = ws_lr.cell(r, 2).value
            print(f'  row{r} Group={grp} 全行: {row}')

# 4) 跟同组其他大转盘对比 — 看 319 的完整奖励池
print('\n【RewardGroup=319 大转盘完整奖励池】')
if 'ActvLuckyWheelReward' in wb_lw.sheetnames:
    ws_lr = wb_lw['ActvLuckyWheelReward']
    # load item names
    wb_it2 = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
    ws_it2 = wb_it2['Item']
    item_name = {}
    for row in ws_it2.iter_rows(min_row=7, max_col=3, values_only=True):
        if isinstance(row[0], int): item_name[row[0]] = row[1]
    wb_it2.close()

    print(f'{"ID":>6} {"ItemID":>6} {"道具名":<20} {"数量":>4} {"权重":>4} {"超级":>4} {"排序":>4}')
    for r in range(7, ws_lr.max_row+1):
        if ws_lr.cell(r, 2).value == 319:
            iid = ws_lr.cell(r, 3).value
            nm = item_name.get(iid, '?') if isinstance(iid, int) else '?'
            print(f'{ws_lr.cell(r,1).value!s:>6} {iid!s:>6} {str(nm)[:20]:<20} {ws_lr.cell(r,5).value!s:>4} {ws_lr.cell(r,6).value!s:>4} {ws_lr.cell(r,7).value!s:>4} {ws_lr.cell(r,8).value!s:>4}')
