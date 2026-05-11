import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 探索 ActvScore.xlsx 结构，找 ContentID=718 的关联
wb = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== {sname} max_row={ws.max_row} max_col={ws.max_column} ===')
    # 列头 (row 5/6)
    for r in range(1, 7):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 18))]
        print(f'  row{r}: {row}')
    # 查 718 和 601 的出现
    print(f'  查 718/601/10071801 在本 sheet 的出现:')
    for r in range(7, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            if v in (718, 601, 10071801):
                # 打整行
                rowdata = [ws.cell(r, cc).value for cc in range(1, min(ws.max_column+1, 18))]
                print(f'    row{r} col{c}={v}: {rowdata}')
                break
