import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb_it = load_workbook('C:/X3/Item.xlsx', data_only=True, read_only=True)
ws_it = wb_it['Item']
item_name = {}
for row in ws_it.iter_rows(min_row=7, max_col=5, values_only=True):
    if isinstance(row[0], int): item_name[row[0]] = row[1]
wb_it.close()

wb_rw = load_workbook('C:/X3/Reward.xlsx', data_only=True)
ws_rw = wb_rw['Reward']

# 所有 reward 数据按 RewardID 索引
reward_data = {}
for r in range(7, ws_rw.max_row+1):
    rid = ws_rw.cell(r, 2).value
    if rid is not None:
        reward_data.setdefault(rid, []).append({
            'ItemType': ws_rw.cell(r, 3).value,
            'ItemID': ws_rw.cell(r, 4).value,
            'MinNum': ws_rw.cell(r, 6).value,
            'DropType': ws_rw.cell(r, 8).value,
        })

def print_reward(rid, indent='  '):
    items = reward_data.get(rid, [])
    for it in items:
        iid = it['ItemID']
        nm = item_name.get(iid, '?') if isinstance(iid, int) else '?'
        itype = {1:'道具', 2:'蓝图', 3:'子包'}.get(it['ItemType'], f"类型{it['ItemType']}")
        print(f'{indent}{itype} {iid} [{nm}] x{it["MinNum"]}')
        if it['ItemType'] == 3 and isinstance(iid, int):
            # 递归展开子包
            print_reward(iid, indent + '  ')

# 1) RankID=169 排行榜奖励
print('='*60)
print('RankID=169 排行榜奖励完整内容')
print('='*60)
排名 = [(30481,'第1名'),(30482,'第2名'),(30483,'第3名'),(30484,'第4-5名'),
       (30485,'第6-10名'),(30486,'第11-20名'),(30487,'第21-50名'),(30488,'第51-100名')]
for rid, label in 排名:
    print(f'\n{label} (RewardID={rid}):')
    print_reward(rid)

# 2) OtherRewardGroup=3019 额外奖励
print(f'\n{"="*60}')
print('OtherRewardGroup=3019 累计抽奖额外奖励')
print('='*60)
档位 = [(1001821,'50次'),(1001822,'150次'),(1001823,'300次'),(1001824,'500次'),
       (1001825,'750次'),(1001826,'1000次'),(1001827,'1500次')]
for rid, label in 档位:
    print(f'\n累计{label} (RewardID={rid}):')
    print_reward(rid)
