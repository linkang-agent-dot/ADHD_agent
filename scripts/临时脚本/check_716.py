import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 1) ActvOnline 里 ContentID=716 或 TC=716 的活动
wb = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws = wb['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws.cell(6, c).value
    if en: headers[en] = c

print('【ActvOnline 里 ContentID=716 的活动】')
for r in range(7, ws.max_row+1):
    cid = ws.cell(r, headers['ContentID']).value
    if cid == 716:
        print(f'  row{r} ActvID={ws.cell(r,1).value} 名={ws.cell(r,3).value!r} 备注={ws.cell(r,2).value!r} Type={ws.cell(r,6).value} IsOn={ws.cell(r,7).value} TC={ws.cell(r,8).value}')

print('\n【ActvOnline 里 TimeController=716 的活动】')
for r in range(7, ws.max_row+1):
    tc = ws.cell(r, headers['TimeController']).value
    if tc == 716:
        print(f'  row{r} ActvID={ws.cell(r,1).value} 名={ws.cell(r,3).value!r} Type={ws.cell(r,6).value} IsOn={ws.cell(r,7).value} ContentID={ws.cell(r, headers["ContentID"]).value}')

# 2) TimeCycle 里 716 这行
wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']
print('\n【TimeCycle ID=716 这行】')
for r in range(6, ws_tc.max_row+1):
    if ws_tc.cell(r, 1).value == 716:
        print(f'  row{r} TC={716} IsOn={ws_tc.cell(r,3).value} TT={ws_tc.cell(r,5).value} T={ws_tc.cell(r,6).value!r} 备注={ws_tc.cell(r,2).value!r}')

# 3) ActvScoreMulti 里 716 的阶段榜是什么
wb_sm = load_workbook('C:/X3/ActvScore.xlsx', data_only=True)
ws_sm = wb_sm['ActvScoreMulti']
print('\n【ActvScoreMulti ContentID=716 的阶段配置】')
for r in range(7, ws_sm.max_row+1):
    if ws_sm.cell(r, 3).value == 716:
        print(f'  row{r} ID={ws_sm.cell(r,1).value} Stage={ws_sm.cell(r,5).value} 阶段名={ws_sm.cell(r,4).value!r} RankID={ws_sm.cell(r,9).value} 备注={ws_sm.cell(r,2).value!r}')
