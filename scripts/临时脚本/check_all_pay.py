import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws_ao.cell(6, c).value
    if en: headers[en] = c

wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']
import re
tc_info = {}
for r in range(6, ws_tc.max_row+1):
    v = ws_tc.cell(r, 1).value
    if v is None: continue
    ison = ws_tc.cell(r, 3).value
    if ison != 1: continue
    tt = ws_tc.cell(r, 5).value
    t_str = ws_tc.cell(r, 6).value or ''
    dt = ws_tc.cell(r, 7).value
    d_str = ws_tc.cell(r, 8).value or ''
    remark = ws_tc.cell(r, 2).value or ''
    start_day = None
    duration_days = None
    if tt == 2:
        m = re.match(r'(\d+)d', str(t_str))
        if m: start_day = int(m.group(1))
        if start_day is None and '00:00:00' in str(t_str) and 'd' not in str(t_str):
            start_day = 0
    if dt == 1 and d_str:
        m2 = re.match(r'(\d+)d', str(d_str))
        if m2: duration_days = int(m2.group(1)) + 1
        if duration_days is None and 'h' in str(d_str):
            duration_days = 1
    tc_info[int(v)] = {'tt':tt,'start_day':start_day,'duration':duration_days,'remark':remark}

# 收集所有 D5-D35 的活动，标记付费类型
activities = []
for r in range(7, ws_ao.max_row+1):
    ison = ws_ao.cell(r, headers['IsOn']).value
    if ison != 1: continue
    aid = ws_ao.cell(r, headers['ID']).value
    tc_id = ws_ao.cell(r, headers['TimeController']).value
    name = ws_ao.cell(r, headers.get('ActvName',3)).value or ''
    atype = ws_ao.cell(r, headers['ActvType']).value
    remark = ws_ao.cell(r, 2).value or ''
    chain = ws_ao.cell(r, headers.get('ChainPackID',32)).value
    recharge = ws_ao.cell(r, headers.get('ActvRechargePoints',47)).value

    if tc_id and isinstance(tc_id, (int, float)):
        tc_id = int(tc_id)
        tc = tc_info.get(tc_id)
        if tc and tc['tt'] == 2 and tc['start_day'] is not None:
            start = tc['start_day']
            dur = tc['duration'] or 999
            end = start + dur - 1

            # 判断付费类型
            pay_type = []
            if chain: pay_type.append(f'链式礼包({chain})')
            if recharge: pay_type.append('累充积分')
            if atype == 10: pay_type.append('大转盘(直接付费)')
            if atype == 11: pay_type.append('BP通行证(直接付费)')
            if atype == 56: pay_type.append('拜访礼包(直接付费)')
            if atype == 50: pay_type.append('许愿池(直接付费)')
            if atype == 60: pay_type.append('海妖转盘(直接付费)')
            if atype == 5 and '抽卡' in remark: pay_type.append('英雄抽卡(消耗钻石)')
            if atype == 5 and '累充' in remark: pay_type.append('累充(直接付费)')
            if atype == 55: pay_type.append('链式礼包(直接付费)')
            if not pay_type:
                # 检查是否有排行榜(间接付费)
                rank = ws_ao.cell(r, headers.get('RankID',21)).value
                if rank: pay_type.append(f'排行榜(间接)')
                else: pay_type.append('无付费点')

            if start <= 35:
                activities.append({
                    'aid':aid,'name':name,'type':atype,'remark':remark,
                    'start':start,'end':min(end,40),'dur':dur,
                    'pay_type':'|'.join(pay_type),'has_direct_pay':any('直接' in p or '链式' in p or '钻石' in p for p in pay_type),
                    'chain':chain,'recharge':recharge
                })

activities.sort(key=lambda x: (x['start'], x['end']))

# 甘特图式输出
print('='*120)
print('D5-D35 活动排期 + 付费属性（🔴直接付费 🟡间接/消耗 🟢无付费）')
print('='*120)
print(f'{"D开始":>4}-{"D结束":>4} {"活动名":<14} {"Type":>4} {"付费类型":<30} {"备注":<20}')

for a in activities:
    if a['start'] >= 5 and a['start'] <= 35:
        end_str = str(a['end']) if a['dur'] < 999 else '∞'
        icon = '🔴' if a['has_direct_pay'] else '🟡' if '间接' in a['pay_type'] or '消耗' in a['pay_type'] else '🟢'
        print(f'D{a["start"]:>3}-D{end_str:>3} {icon} {str(a["name"])[:14]:<14} {a["type"]:>4} {a["pay_type"][:30]:<30} {str(a["remark"])[:20]:<20}')

# D14-D35 每天的"直接付费活动数"
print(f'\n{"="*120}')
print('D14-D35 每天直接付费活动数量')
print(f'{"="*120}')
for d in range(14, 36):
    direct = [a for a in activities if a['start'] <= d <= a['end'] and a['has_direct_pay']]
    indirect = [a for a in activities if a['start'] <= d <= a['end'] and not a['has_direct_pay'] and '间接' in a['pay_type']]
    free = [a for a in activities if a['start'] <= d <= a['end'] and '无付费' in a['pay_type']]

    d_bar = '🔴' * len(direct)
    i_bar = '🟡' * len(indirect)
    names = ', '.join(f'{a["name"][:8]}' for a in direct)
    print(f'D{d:>2}: 直接{len(direct):>2}个{d_bar:<16} 间接{len(indirect)}个{i_bar:<6} | {names}')
