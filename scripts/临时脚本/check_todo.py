import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

# 加载配置表
wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws_ao.cell(6, c).value
    if en: headers[en] = c

wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']

# 尼罗河10个活动当前配置状态
nile = [101023, 101024, 102234, 105602, 101825, 105010, 10071801, 101334, 101335, 100594]
act_names = {101023:'黄金卷轴(大转盘)',101024:'折扣券大转盘',102234:'圣甲虫试炼(BP)',
             105602:'金字塔之城(拜访)',101825:'象形密文(拼图)',105010:'许愿池',
             10071801:'蓝莲花宴(酒馆)',101334:'方尖碑集市(兑换)',101335:'往昔盛典',100594:'黄金沙漏(累充)'}

# 目标方案
plan = {
    101023: {'action':'D20开','tc':'1825','need':'TimeController改为1825 + ActvScoreMulti已改本服rank'},
    102234: {'action':'D20开','tc':'1825','need':'已OK(TC=1825)'},
    105602: {'action':'D24开','tc':'新建D24','need':'新建TC + TimeController指向新TC'},
    101825: {'action':'D20开','tc':'1825','need':'已OK(TC=1825)'},
    105010: {'action':'不上线','tc':'-','need':'跨服限制'},
    10071801: {'action':'D20开','tc':'1825','need':'TimeController改为1825 + 排行榜改本服 + ActvScoreMulti已改'},
    101334: {'action':'D20开','tc':'1825','need':'TimeController改为1825'},
    101335: {'action':'不上线','tc':'-','need':'从未配置'},
    100594: {'action':'D20开','tc':'1825','need':'已OK(TC=1825)'},
    101024: {'action':'不上线','tc':'-','need':'本期不上'},
}

print('='*120)
print('尼罗河常态化改造 — 待办清单')
print('='*120)

issues = []

for aid in nile:
    for r in range(7, ws_ao.max_row+1):
        if ws_ao.cell(r, 1).value != aid: continue
        name = act_names[aid]
        tc_id = ws_ao.cell(r, headers['TimeController']).value
        atype = ws_ao.cell(r, headers['ActvType']).value
        ison = ws_ao.cell(r, headers['IsOn']).value
        csr = ws_ao.cell(r, headers.get('CrossServerRank',19)).value
        rank_id = ws_ao.cell(r, headers.get('RankID',21)).value
        p = plan[aid]

        print(f'\n--- {aid} {name} ---')
        print(f'  目标: {p["action"]}')
        print(f'  当前 TC={tc_id}, Type={atype}, IsOn={ison}, CSR={csr}, RankID={rank_id}')

        # 检查问题
        if p['action'] == '不上线':
            print(f'  -> 不上线，无需改动')
            continue

        # TC 检查
        if p['action'] == 'D20开':
            if tc_id != 1825:
                issues.append((aid, name, 'ActvOnline', f'TimeController: {tc_id} -> 1825'))
                print(f'  !! ActvOnline.TimeController 需改: {tc_id} -> 1825')
            else:
                print(f'  OK TC=1825')
        elif p['action'] == 'D24开':
            issues.append((aid, name, 'TimeCycle+ActvOnline', f'新建TC(D24 TT=2 23d) + TimeController指向新TC'))
            print(f'  !! 需新建 TimeCycle 行 + 改 TimeController')

        # 跨服排行检查
        if csr == 1 and p['action'] != '不上线':
            issues.append((aid, name, 'ActvOnline', f'CrossServerRank: 1 -> 空 (改本服)'))
            print(f'  !! CrossServerRank=1 需改为空')

        # RankID 检查 (跨服榜->本服榜)
        if rank_id and isinstance(rank_id, (int, float)):
            rank_id = int(rank_id)
            # 检查是否跨服榜
            wb_rk = load_workbook('C:/X3/Rank.xlsx', data_only=True)
            ws_rk = wb_rk['RankCfg']
            for rr in range(6, ws_rk.max_row+1):
                if ws_rk.cell(rr, 1).value == rank_id:
                    rt = ws_rk.cell(rr, 4).value  # RankType
                    remark = ws_rk.cell(rr, 2).value
                    if rt == 12:  # 跨服
                        issues.append((aid, name, 'ActvOnline+Rank', f'RankID={rank_id}是跨服(RankType=12), 需改本服'))
                        print(f'  !! RankID={rank_id} 是跨服榜({remark}), 需改为本服')
                    else:
                        print(f'  OK RankID={rank_id} 是本服(RankType={rt})')
                    break
        break

# 汇总
print(f'\n{"="*120}')
print(f'待改动汇总 ({len(issues)} 项)')
print(f'{"="*120}')

# 按表分组
from collections import defaultdict
by_table = defaultdict(list)
for aid, name, table, desc in issues:
    by_table[table].append((aid, name, desc))

for table, items in sorted(by_table.items()):
    print(f'\n  【{table}】')
    for aid, name, desc in items:
        print(f'    {aid} {name}: {desc}')
EOF
