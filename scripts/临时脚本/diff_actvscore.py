import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

def scan_all(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    result = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = {}
        for r in range(1, ws.max_row+1):
            row_tuple = tuple(ws.cell(r, c).value for c in range(1, ws.max_column+1))
            rows[r] = row_tuple
        result[sname] = rows
    wb.close()
    return result

old = scan_all('C:/Users/linkang/x3_pristine/ActvScore_before_A_fix.xlsx')
new = scan_all('C:/X3/ActvScore.xlsx')

print('【ActvScore.xlsx 变更对比：备份 vs 当前】')
for sname in set(old) | set(new):
    o = old.get(sname, {})
    n = new.get(sname, {})
    all_rows = set(o) | set(n)
    changes = []
    for r in sorted(all_rows):
        orow = o.get(r, ())
        nrow = n.get(r, ())
        if orow != nrow:
            changes.append((r, orow, nrow))
    if changes:
        print(f'\n--- Sheet: {sname} ({len(changes)} 行变更) ---')
        for r, orow, nrow in changes:
            print(f'  row{r}:')
            max_c = max(len(orow), len(nrow))
            for c in range(max_c):
                ov = orow[c] if c < len(orow) else None
                nv = nrow[c] if c < len(nrow) else None
                if ov != nv:
                    print(f'    col{c+1}: {ov!r} → {nv!r}')
