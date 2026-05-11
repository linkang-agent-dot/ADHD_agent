import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook

wb = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws = wb['ActvOnline']

# 扫描所有列头里含 rank/排行/榜 的
print('【所有含 rank/排行/榜/multi 的列】')
for c in range(1, 60):
    zh = ws.cell(5, c).value
    en = ws.cell(6, c).value
    note = ws.cell(4, c).value
    combo = ' '.join(str(x) for x in [zh, en, note] if x)
    if 'rank' in combo.lower() or '排行' in combo or '榜' in combo or 'multi' in combo.lower() or '排名' in combo:
        print(f'  col{c}: 中={zh!r} en={en!r}')
        if note: print(f'           note={str(note)[:120]!r}')

# 找 10071801 的行，打印所有非空列
print('\n【10071801 蓝莲花宴 所有非空列】')
for r in range(7, ws.max_row+1):
    if ws.cell(r, 1).value == 10071801:
        for c in range(1, 60):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                zh = ws.cell(5, c).value
                en = ws.cell(6, c).value
                print(f'  col{c} ({zh}/{en}) = {v!r}')
        break

# 也看 100701（另一个 Type=7 最佳酒馆活动）作为对照
print('\n【对照：100701 最佳酒馆 所有非空列】')
for r in range(7, ws.max_row+1):
    if ws.cell(r, 1).value == 100701:
        for c in range(1, 60):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                zh = ws.cell(5, c).value
                en = ws.cell(6, c).value
                print(f'  col{c} ({zh}/{en}) = {v!r}')
        break

wb.close()
