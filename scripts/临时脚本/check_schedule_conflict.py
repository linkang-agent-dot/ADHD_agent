import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl.worksheet.table as tbl_mod
from openpyxl.descriptors.base import String
tbl_mod.Table.ref = String(allow_none=False)
from openpyxl import load_workbook
import re

# 加载 TimeCycle
wb_tc = load_workbook('C:/X3/TimeCycle.xlsx', data_only=True)
ws_tc = wb_tc['TimeCycle']
tc_info = {}
for r in range(6, ws_tc.max_row+1):
    v = ws_tc.cell(r, 1).value
    if v is None: continue
    ison = ws_tc.cell(r, 3).value
    if ison != 1: continue
    tt = ws_tc.cell(r, 5).value
    t_str = ws_tc.cell(r, 6).value or ''
    dt = ws_tc.cell(r, 7).value
    d_str = ws_tc.cell(r, 8).value or ''
    remark = ws_tc.cell(r, 2).value or ''

    # 解析开服时间类型(TT=2)的开始天数和持续天数
    start_day = None
    duration_days = None

    if tt == 2:  # 开服时间
        # 解析 "20d 00:00:00" → 20天
        m = re.match(r'(\d+)d', str(t_str))
        if m: start_day = int(m.group(1))
        # 也处理 "00:00:00" (当天=D0)
        if start_day is None and '00:00:00' in str(t_str) and 'd' not in str(t_str):
            start_day = 0

    if dt == 1 and d_str:  # 持续一段时间
        # 解析 "13d23h59m59s" → ~14天
        m2 = re.match(r'(\d+)d', str(d_str))
        if m2: duration_days = int(m2.group(1)) + 1
        # "23h59m59s" = ~1天
        if duration_days is None and 'h' in str(d_str):
            duration_days = 1

    tc_info[int(v)] = {
        'tt': tt, 'start_day': start_day, 'duration': duration_days,
        'remark': remark, 't_str': t_str, 'd_str': d_str
    }

# 加载 ActvOnline
wb_ao = load_workbook('C:/X3/ActvOnline.xlsx', data_only=True)
ws_ao = wb_ao['ActvOnline']
headers = {}
for c in range(1, 60):
    en = ws_ao.cell(6, c).value
    if en: headers[en] = c

# 收集所有 IsOn=1 且 TC 为开服时间(TT=2)的活动
activities = []
for r in range(7, ws_ao.max_row+1):
    ison = ws_ao.cell(r, headers['IsOn']).value
    if ison != 1: continue
    aid = ws_ao.cell(r, headers['ID']).value
    tc_id = ws_ao.cell(r, headers['TimeController']).value
    name = ws_ao.cell(r, headers.get('ActvName', 3)).value
    atype = ws_ao.cell(r, headers['ActvType']).value
    remark = ws_ao.cell(r, 2).value or ''

    if tc_id and isinstance(tc_id, (int, float)):
        tc_id = int(tc_id)
        tc = tc_info.get(tc_id)
        if tc and tc['tt'] == 2 and tc['start_day'] is not None:
            start = tc['start_day']
            dur = tc['duration'] or 999  # 无限持续
            end = start + dur - 1
            activities.append({
                'aid': aid, 'name': name, 'type': atype, 'remark': remark,
                'tc': tc_id, 'start': start, 'end': end, 'dur': dur,
                'tc_remark': tc['remark']
            })

# 按开始天数排序
activities.sort(key=lambda x: (x['start'], x['end']))

# 输出 D0-D40 范围内的活动时间线
print('='*100)
print('X3 开服生命周期活动排期（TT=2 开服时间，IsOn=1）— D0~D40')
print('='*100)
print(f'{"开始":>4}-{"结束":>4} {"持续":>4} {"ActvID":>10} {"Type":>4} {"活动名":<14} {"备注":<25} {"TC备注":<30}')
for a in activities:
    if a['start'] <= 40:
        end_str = str(a['end']) if a['dur'] < 999 else '∞'
        print(f'D{a["start"]:>3}-D{end_str:>3} {a["dur"]:>4}天 {a["aid"]:>10} {a["type"]:>4} {str(a["name"])[:14]:<14} {str(a["remark"])[:25]:<25} {str(a["tc_remark"])[:30]:<30}')

# 分析 D16-D30 每天有多少活动在跑
print(f'\n{"="*100}')
print('D14-D35 每天活跃活动数量 + 活动列表')
print(f'{"="*100}')
for d in range(14, 36):
    active = [a for a in activities if a['start'] <= d <= a['end']]
    names = ', '.join(f'{a["name"]}({a["remark"][:8]})' for a in active[:8])
    bar = '█' * len(active)
    print(f'D{d:>2}: {len(active):>2}个 {bar:<20} {names}')

# 4个候选方案的冲突分析
print(f'\n{"="*100}')
print('4个候选方案的排期冲突分析')
print(f'{"="*100}')
candidates = [(16,'S',14),(18,'A',14),(21,'B',10),(24,'C',14)]
for start, grade, dur in candidates:
    end = start + dur - 1
    # 统计期间有多少活动同时在跑
    overlap = set()
    for d in range(start, end+1):
        for a in activities:
            if a['start'] <= d <= a['end']:
                overlap.add(a['aid'])

    # 其中哪些是节日/活动类型（非常驻功能）
    festival = [a for a in activities if a['aid'] in overlap and a['start'] >= 5 and a['dur'] < 100]
    permanent = [a for a in activities if a['aid'] in overlap and (a['dur'] >= 100 or a['start'] < 5)]

    print(f'\n  方案{grade} D{start}→D{end}:')
    print(f'    期间共 {len(overlap)} 个活动同时在跑')
    print(f'    其中限时活动 {len(festival)} 个:')
    for a in sorted(festival, key=lambda x: x['start']):
        print(f'      D{a["start"]}-D{a["end"]} {a["name"]} ({a["remark"][:20]})')
    print(f'    常驻/长期活动 {len(permanent)} 个（不构成冲突）')
