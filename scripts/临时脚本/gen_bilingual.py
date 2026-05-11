import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = Workbook()
HDR_BG="1F4E79"; SEC_BG="D6E4F0"; SUB_BG="EBF5FB"; NOTE_BG="FFF2CC"; GREEN_BG="E2EFDA"; CN_BG="EBF5FB"; EN_BG="E8F5E9"

def s(ws,coord,value,bold=False,size=11,fg="FF000000",bg=None,wrap=True,ha="left",va="top"):
    c=ws[coord]; c.value=value
    c.font=Font(name="Arial",size=size,bold=bold,color=fg)
    c.alignment=Alignment(wrap_text=wrap,horizontal=ha,vertical=va)
    if bg: c.fill=PatternFill("solid",fgColor=bg)
    return c
def hdr(ws,coord,val,size=11): s(ws,coord,val,bold=True,size=size,fg="FFFFFF",bg=HDR_BG,ha="center",va="center")
def sec(ws,coord,val): s(ws,coord,val,bold=True,size=11,fg="1F4E79",bg=SEC_BG,ha="left",va="center")
def sub(ws,coord,val): s(ws,coord,val,bold=True,size=10,fg="1F4E79",bg=SUB_BG,ha="left",va="center")
def cel(ws,coord,val,bg="FFFFFF",bold=False,ha="left"): s(ws,coord,val,bold=bold,size=10,bg=bg,ha=ha,va="top")
def note(ws,coord,val): s(ws,coord,val,size=9,fg="7F6000",bg=NOTE_BG,ha="left",va="top")
def col(n): return get_column_letter(n)
